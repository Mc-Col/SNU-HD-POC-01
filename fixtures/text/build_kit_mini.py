# -*- coding: utf-8 -*-
"""fixtures/text/kit_mini.xlsx — 채점 도구용 최소 골든셋. 합성 데이터.

  python fixtures/text/build_kit_mini.py

실제 labeling_kit.xlsx 의 '라벨링' 시트 구조만 흉내 낸다.
  1행 = 필드명(정답값·원문라벨 두 칸을 덮는다) / 2행 = 소제목 / 3행부터 문서
"""
import os

import openpyxl

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "kit_mini.xlsx")

META = ["문서ID", "파일명", "포맷", "연식", "문서분류", "총페이지", "사양표 페이지",
        "라벨러", "소요(분)"]
FIELDS = ["ENGINEERING TAG NO.", "MANUFACTURER", "MODEL NO.",
          "ACTUATOR FAIL ACTION", "RATED CV"]        # RATED CV = 스키마에 없는 이름

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "라벨링"

for i, m in enumerate(META, start=1):
    ws.cell(1, i).value = "문서 정보"
    ws.cell(2, i).value = m
for j, name in enumerate(FIELDS):
    c = 10 + j * 2
    ws.cell(1, c).value = name
    ws.cell(2, c).value = "정답값"
    ws.cell(2, c + 1).value = "원문라벨"

# 라벨러를 넣는다 — 골든셋에 사람 라벨과 AI 초안이 섞여 있어 채점을 나눠 세야 한다
# (2026-08-26). 사람이 만든 정답만으로 잰 숫자가 발표에 쓰는 값이다.
ROWS = [
    # 파일명, 문서분류, 사양표 시트, 라벨러, [(정답값, 원문라벨)]
    ("excel_layouts.xlsx", "datasheet", 1, "사람", [
        ("44-LV-999", "Tag No."), ("FISHER", "(하단 꼬리말)"),
        ("880-2221", "Model No."), ("FAIL CLOSE", "Fail Position"), ("N/A", "")]),
    ("pdf_basic.pdf", "datasheet", 1, "AI초안(Claude)", [
        ("11-FV-999", "Tag"), ("FISHER", "Manufacturer"),
        ("667-ED", "Model No."), ("FAIL CLOSE", "Fail/Air-To"), ("N/A", "")]),
    ("excel_basic.xlsx", "out_of_scope", None, "사람", [("", "")] * 5),
]
for i, (fn, cls, page, who, pairs) in enumerate(ROWS):
    r = 3 + i
    ws.cell(r, 1).value = f"d{i + 1:03d}"
    ws.cell(r, 2).value = fn
    ws.cell(r, 5).value = cls
    ws.cell(r, 7).value = page
    ws.cell(r, 8).value = who
    for j, (val, lab) in enumerate(pairs):
        ws.cell(r, 10 + j * 2).value = val or None
        ws.cell(r, 11 + j * 2).value = lab or None

wb.save(OUT)
print("생성:", OUT)
