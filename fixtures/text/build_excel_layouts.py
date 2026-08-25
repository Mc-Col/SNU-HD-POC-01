# -*- coding: utf-8 -*-
"""fixtures/text/excel_layouts.xlsx — 실물에서 관찰된 엑셀 배치. 합성 데이터.

  python fixtures/text/build_excel_layouts.py

  [SPEC] 44LV001 계열 — 구역 라벨 A열 · 항목 라벨 B열 · 값 E열 (3칸 간격)
  [TEST] 11FV048 계열 — 라벨이 병합(B:H)되고 값이 I열, 오른쪽에 같은 구조가 한 번 더
"""
import os

import openpyxl

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel_layouts.xlsx")

wb = openpyxl.Workbook()

ws = wb.active
ws.title = "SPEC"
ws["A1"] = "CONTROL VALVE SPECIFICATIONS"
ws["A6"], ws["C6"] = "Tag No.", "44-LV-999"
ws["B9"], ws["E9"] = "Model No.", "880-2221"
ws["A10"], ws["B10"], ws["E10"] = "BODY", "Body Type", "GLOBE"          # 미매핑(유사표현 없음)
# 구역 이름표는 실물처럼 세로로 병합한다 (44LV001 은 A열에 BODY·MATERIAL 이 세로로 선다).
# 병합돼 있어야 파서가 이름표로 알아보고 라벨로 읽지 않는다.
ws["B11"], ws["E11"] = "Rating", "ANSI CLASS 300"                        # 미매핑
ws["B18"], ws["E18"] = "Characteristic", "LINEAR"
ws["A22"], ws["B22"], ws["E22"] = "MATERIAL", "Body", "ASTM A216 Gr-WCB"  # 미매핑
ws.merge_cells("A10:A21")          # BODY 구역
ws.merge_cells("A22:A32")          # MATERIAL 구역
ws["B35"], ws["E35"] = "Fail Position", "CLOSE"
ws["B36"], ws["E36"] = "Actuator Type", "DIAPHRAGM"
ws["B42"], ws["E42"] = "Positioner Type", "PNEUMATIC"

ws2 = wb.create_sheet("TEST")
ws2["A1"] = "CONTROL VALVE RETROFIT"
# 라벨이 B:H 로 병합되고 값은 I열 — 병합을 못 따라가면 7칸 밖이라 놓친다
for row, (lab, val, lab2, val2) in {
    5:  ("TAG NO", "11-FV-999", None, None),
    7:  ("Body Model(Type)", "657-ED(GLOBE)", "Actuator Type", "DIAPHRAGM"),
    9:  ("Leakage Spec.", "CLASS 2", "Fail Position", "OPEN"),
    17: ("Trim Form", "CAGE-BALANCED", "Maker", "FISHER"),
}.items():
    ws2.cell(row, 2).value = lab
    ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    ws2.cell(row, 9).value = val
    if lab2:
        ws2.cell(row, 19).value = lab2
        ws2.merge_cells(start_row=row, start_column=19, end_row=row, end_column=25)
        ws2.cell(row, 26).value = val2

wb.save(OUT)
print("생성:", OUT)
