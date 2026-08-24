# -*- coding: utf-8 -*-
"""fixtures/text/excel_basic.xlsx 생성 — 합성 데이터. 회사 문서를 넣지 않는다.

  python fixtures/text/build_excel_basic.py

raw_file/ 은 Git 에서 제외되므로 fixture 에 실물 문서를 쓸 수 없다.
대신 실제 양식에서 관찰된 배치만 흉내 낸 최소 시트를 만든다.
"""
import os

import openpyxl

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel_basic.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SPEC"

ws["A1"] = "CONTROL VALVE SPECIFICATION"      # 병합 제목 — 라벨 아님
ws.merge_cells("A1:D1")

# 라벨 좌 · 값 우 (가장 흔한 배치)
ws["A3"], ws["B3"] = "Tag", "11-FV-999"                       # 유사표현
ws["A4"], ws["B4"] = "Manufacturer", "FISHER"                 # 표준명
ws["A5"], ws["B5"] = "Model No.", "667-ED"                    # 표준명
ws["A6"], ws["B6"] = "Fail Position", "Air Fails Valve to Close"   # 유사표현
ws["A7"], ws["B7"] = "Application", "RECYCLE TO DHC FEED FILTERS"  # 유사표현
ws["A8"], ws["B8"] = "Spring Range", "0.4 - 2.0 bar"          # 스키마에 없는 항목
ws["A9"], ws["B9"] = "Body Size", "4 IN"                      # 유사표현 미등록 → 미매핑

# 같은 행 오른쪽 블록
ws["D3"], ws["E3"] = "Characteristic", "EQUAL %"

# 병합 라벨 + 아래 값
ws["A11"] = "VALVE BODY SIZE"
ws.merge_cells("A11:B11")
ws["A12"] = "4 IN"

# 라벨은 있으나 값이 비어 있음
ws["A14"] = "Viscosity"

# 두 번째 시트 — 먼저 찾은 값이 이겨야 한다
ws2 = wb.create_sheet("PHOTO")
ws2["A1"], ws2["B1"] = "Manufacturer", "MASONEILAN"

wb.save(OUT)
print("생성:", OUT)
