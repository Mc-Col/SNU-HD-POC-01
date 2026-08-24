# -*- coding: utf-8 -*-
"""골든셋(라벨링 킷) 읽기 — 평가의 입력

    from eval.kit import read_kit

    for row in read_kit("readme/labeling_kit.xlsx"):
        row.doc_id, row.file, row.spec_page, row.truth["rated_cv"]

킷은 특정 모듈의 것이 아니라 **평가 전체의 입력**이므로 여기 둔다.
`src/parsers/text/score_against_kit.py` 도 자체 리더를 갖고 있는데, 그쪽은
텍스트 파서만 재는 도구다(그 파일 docstring 참조). 리더를 합치려면
이쪽으로 모으는 것이 맞다 — 킷의 구조를 아는 곳이 하나여야 한다.

킷 구조 (tools/gen_labelkit.py 가 만든다)
    1행  그룹명 (병합) — 필드 표준명 또는 "문서 정보" / "비고"
    2행  하위 헤더 — "정답값" / "원문라벨", 문서정보는 항목명
    3행~ 문서 1건당 1행
"""
from __future__ import annotations

import os
import re
from dataclasses import dataclass, field as dc_field

SHEET = "라벨링"

# 문서 정보 열 이름 → 속성
INFO = {
    "문서ID": "doc_id", "파일명": "file", "포맷": "fmt", "연식": "vintage",
    "문서분류": "doc_class", "총페이지": "pages", "사양표 페이지": "spec_page",
    "라벨러": "labeler", "소요(분)": "minutes",
}


@dataclass
class KitRow:
    doc_id: str = ""
    file: str = ""
    fmt: str = ""
    vintage: str = ""
    doc_class: str = ""
    pages: int | None = None
    spec_page: int | None = None
    labeler: str = ""
    minutes: str = ""
    note: str = ""
    truth: dict[str, str] = dc_field(default_factory=dict)      # field_key → 정답값
    raw_label: dict[str, str] = dc_field(default_factory=dict)  # field_key → 원문라벨
    path: str | None = None                                      # 실제 파일 경로

    @property
    def uncertain_fields(self) -> tuple[str, ...]:
        """라벨러가 물음표를 붙인 필드 — 집계에서 따로 센다."""
        return tuple(k for k, v in self.truth.items()
                     if str(v or "").strip().startswith("?"))


def _int(v) -> int | None:
    try:
        return int(str(v).strip())
    except (TypeError, ValueError):
        return None


def _norm_name(s: str) -> str:
    """킷의 그룹명을 필드 표준명으로. 확장 필드의 `· ` 접두어와 `※안전` 을 뗀다."""
    t = re.sub(r"※.*$", "", str(s or "")).strip()
    return t.lstrip("·").strip()


def read_kit(path: str, schema_mod=None) -> list[KitRow]:
    """킷을 읽어 채워진 행만 돌려준다. 필드명은 `schema` 로 key 로 바꾼다.

    스키마에 없는 킷 컬럼은 조용히 버리지 않고 `unmatched` 로 알린다 —
    표준이 바뀌면 여기서 먼저 드러나야 한다(철학 5).
    """
    import openpyxl
    if schema_mod is None:
        from src import schema as schema_mod

    by_name = {schema_mod.norm_label(f.name): f.key for f in schema_mod.all_fields()}

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET]
    nc = ws.max_column

    g1 = [ws.cell(1, c).value for c in range(1, nc + 1)]
    g2 = [ws.cell(2, c).value for c in range(1, nc + 1)]
    cur, group = None, []
    for x in g1:
        if x:
            cur = x
        group.append(cur)

    info_col, val_col, lab_col, note_col = {}, {}, {}, None
    unmatched = []
    for c in range(nc):
        grp, sub = group[c], str(g2[c] or "").strip()
        if grp == "문서 정보" and sub in INFO:
            info_col[INFO[sub]] = c
        elif grp == "비고":
            note_col = c
        elif sub in ("정답값", "원문라벨"):
            key = by_name.get(schema_mod.norm_label(_norm_name(grp)))
            if key is None:
                if _norm_name(grp) not in unmatched:
                    unmatched.append(_norm_name(grp))
                continue
            (val_col if sub == "정답값" else lab_col)[key] = c

    rows = []
    for r in range(3, ws.max_row + 1):
        cells = [ws.cell(r, c + 1).value for c in range(nc)]
        if sum(1 for v in cells if v not in (None, "")) <= 1:
            continue
        row = KitRow()
        for attr, c in info_col.items():
            v = cells[c]
            setattr(row, attr, _int(v) if attr in ("pages", "spec_page")
                    else ("" if v is None else str(v).strip()))
        if note_col is not None and cells[note_col]:
            row.note = str(cells[note_col]).strip()
        for k, c in val_col.items():
            if cells[c] not in (None, ""):
                row.truth[k] = str(cells[c]).strip()
        for k, c in lab_col.items():
            if cells[c] not in (None, ""):
                row.raw_label[k] = str(cells[c]).strip()
        rows.append(row)

    read_kit.unmatched = unmatched      # 호출자가 보고할 수 있게 남긴다
    return rows


read_kit.unmatched = []


def locate(rows: list[KitRow], root: str = "raw_file") -> list[str]:
    """각 행의 실제 파일을 찾아 `row.path` 를 채운다. → 못 찾은 파일명 목록.

    조원 배분 폴더(`raw_file/<이름>/`)에 있는 것도 찾는다.
    """
    index = {}
    for dirpath, _dirs, files in os.walk(root):
        for f in files:
            index.setdefault(f.lower(), os.path.join(dirpath, f))
    missing = []
    for row in rows:
        row.path = index.get((row.file or "").lower())
        if row.path is None and row.file:
            missing.append(row.file)
    return missing
