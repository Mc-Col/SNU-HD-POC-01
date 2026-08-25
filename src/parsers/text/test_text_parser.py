# -*- coding: utf-8 -*-
"""fixtures/text/ 로 자기 검증한다. 남의 모듈을 기다리지 않는다."""
from __future__ import annotations

import json
import os
import sys

import pytest

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from src.contracts import ParserType                       # noqa: E402
from src.parsers.text import FieldIndex, normalize_label, parse_excel   # noqa: E402

FIXTURE = os.path.join(ROOT, "fixtures", "text", "excel_basic.xlsx")
EXPECTED = os.path.join(ROOT, "fixtures", "text", "excel_basic.expected.json")


@pytest.fixture(scope="module")
def expected() -> dict:
    with open(EXPECTED, encoding="utf-8") as f:
        return json.load(f)


@pytest.fixture(scope="module")
def parsed():
    return parse_excel(FIXTURE)


# ── 라벨 정규화 ────────────────────────────────────────────────
@pytest.mark.parametrize("raw,norm", [
    ("Model No.", "MODELNO"),
    ("MODEL NO", "MODELNO"),
    ("  model  no.  ", "MODELNO"),
    ("Req'd Flow Coeff., Cv", "REQDFLOWCOEFFCV"),
    (None, ""),
])
def test_라벨_정규화는_표기_흔들림을_흡수한다(raw, norm):
    assert normalize_label(raw) == norm


def test_유사표현은_스키마에서만_읽는다():
    ix = FieldIndex.load()
    # 필드 수를 숫자로 박지 않는다. 표준은 실물 라벨링을 거치며 바뀐다
    # (30 → 28: C027 에서 POSITIONER TYPE·MIN/MAX TEMP 삭제, RATED CV MAX 병합).
    # 스키마가 담은 수와 같은지만 본다 — FieldIndex 가 조용히 빠뜨리지 않는지 확인.
    from src import schema as _schema
    assert ix.field_count == len(_schema.all_fields())
    assert ix.lookup("Tag").key == "engineering_tag_no"
    assert ix.lookup("Fail Position").key == "actuator_fail_action"


def test_여러_필드에_걸린_표기는_구역_없이_매핑하지_않는다():
    """`Maker` · `Model` 처럼 부품마다 되풀이되는 표기는 일부러 여러 필드에 등록한다.

    구역(section)을 읽으면 갈라 쓰고, 못 읽으면 아무것도 만들지 않는다.
    이름의 주인(표준명)만 예외로 이긴다 — `MODEL NO.` 는 MODEL NO. 의 이름이다.
    """
    ix = FieldIndex.load()
    assert ix.collisions, "구역으로 갈리는 표기가 하나도 없다면 사전이 낡은 것이다"
    for label, _, _ in ix.collisions:
        hit = ix.lookup(label)
        assert hit is None or hit.matched_on == "name", (
            f"{label!r} 가 구역 없이 {hit and hit.key} 로 매핑된다")


# ── 파서 출력 ──────────────────────────────────────────────────
def test_기대출력과_정확히_일치한다(parsed, expected):
    got = [
        {
            "field_key": r.field_key,
            "raw_value": r.raw_value,
            "raw_label": r.raw_label,
            "source_locator": r.source_locator,
            "page": r.page,
            "confidence": r.confidence,
            "note": r.note,
        }
        for r in parsed.records
    ]
    assert got == expected["records"]


def test_미매핑_라벨을_삼키지_않는다(parsed, expected):
    got = [
        {"text": u.text, "source_locator": u.source_locator,
         "neighbor_value": u.neighbor_value}
        for u in parsed.unmapped
    ]
    assert got == expected["unmapped"]


def test_raw_label_은_문서_표기_그대로다(parsed):
    """유사표현 사전이 여기서 자란다 — 표준명으로 덮어쓰면 안 된다."""
    by = parsed.by_key()
    assert by["engineering_tag_no"].raw_label == "Tag"
    assert by["actuator_fail_action"].raw_label == "Fail Position"


def test_값이_없으면_만들어내지_않는다(parsed):
    """모르면 None. 추정값을 넣지 않는다."""
    r = parsed.by_key()["viscosity"]
    assert r.raw_value is None
    assert r.found is False
    assert r.note                                   # 사유를 반드시 남긴다


def test_먼저_찾은_값이_이긴다(parsed):
    """PHOTO 시트의 MASONEILAN 이 SPEC 시트 값을 덮어쓰지 않는다."""
    assert parsed.by_key()["manufacturer"].raw_value == "FISHER"


def test_계약_타입을_그대로_반환한다(parsed):
    for r in parsed.records:
        assert r.parser is ParserType.EXCEL
        assert r.source_locator and "!" in r.source_locator


def test_같은_입력이면_같은_출력이다():
    a, b = parse_excel(FIXTURE), parse_excel(FIXTURE)
    assert [r.field_key for r in a.records] == [r.field_key for r in b.records]
    assert [r.source_locator for r in a.records] == [r.source_locator for r in b.records]


def test_복합_라벨을_쪼갠다(parsed):
    """'Body Model(Type)' = '657-ED(GLOBE)' → 모델과 바디 형상."""
    r = parsed.by_key()["model_no"]
    assert r.raw_value == "657-ED"
    assert "복합 라벨" in r.note


def test_구분자형_복합_라벨도_쪼갠다(parsed):
    r = parsed.by_key()["valve_leakage_class"]
    assert r.raw_value == "ANSI IV"          # "120 psi / ANSI IV" 의 뒤 조각
    assert r.raw_label == "Shutoff Class"


def test_스키마에_없는_라벨은_버리지_않고_미매핑으로_남긴다(parsed):
    """`Spring Range` 는 28필드에 없다. 값이 있어도 버리지 않고 드러내야 한다 —
    그래야 유사표현·신규필드 후보가 실물에서 수집된다.

    2026-08-24 — 이 테스트는 원래 `valve_body_type` 을 대상으로 삼았는데
    그 필드가 신설되어 이제 정상 매핑된다(아래 테스트로 옮겼다).
    """
    pending = [u for u in parsed.unmapped if u.text == "Spring Range"]
    assert pending and pending[0].neighbor_value == "0.4 - 2.0 bar"


def test_복합_라벨_조각이_스키마에_있으면_매핑된다(parsed):
    """`Body Model(Type)` = `657-ED(GLOBE)` 의 뒤 조각이 valve_body_type 이다.

    이 필드는 2026-08-24 신설되었다. 그 전에는 미매핑으로 남았다 —
    미매핑 수집이 신규 필드 발견으로 이어진 실제 사례다.
    """
    r = parsed.by_key()["valve_body_type"]
    assert r.raw_value == "GLOBE"
    assert "복합 라벨" in r.note


# ── 실물 배치 (fixtures/text/excel_layouts.xlsx) ───────────────
LAYOUTS = os.path.join(ROOT, "fixtures", "text", "excel_layouts.xlsx")
LAYOUTS_EXPECTED = os.path.join(ROOT, "fixtures", "text", "excel_layouts.expected.json")


@pytest.fixture(scope="module")
def layouts():
    return parse_excel(LAYOUTS)


def test_실물_배치_기대출력과_일치한다(layouts):
    with open(LAYOUTS_EXPECTED, encoding="utf-8") as f:
        expected = json.load(f)
    got = [{k: getattr(r, k) for k in
            ("field_key", "raw_value", "raw_label", "source_locator",
             "page", "confidence", "note")}
           for r in layouts.records]
    assert got == expected["records"]


def test_라벨과_값이_세_칸_떨어져도_읽는다(layouts):
    """44LV001 계열 — 항목 라벨 B열, 값 E열."""
    assert layouts.by_key()["model_no"].source_locator == "SPEC!E9"


def test_병합_라벨_바깥의_값을_읽는다():
    """11FV048 계열 — 라벨이 B:H 로 병합되고 값은 I열 (7칸 밖).

    SCAN_RIGHT 는 4 이지만 병합 범위 끝에서부터 세므로 닿는다.
    """
    by = parse_excel(LAYOUTS, sheets=["TEST"]).by_key()
    assert by["model_no"].raw_value == "657-ED"        # 'Body Model(Type)' 복합 분해
    assert by["model_no"].source_locator == "TEST!I7"
    assert by["actuator_type"].source_locator == "TEST!Z7"


def test_시트를_지정하면_그_시트만_본다():
    """Triage 가 사양표 시트를 알려주면 사진·이력 시트는 보지 않는다."""
    both = parse_excel(LAYOUTS)
    one = parse_excel(LAYOUTS, sheets=[1])
    assert {r.page for r in one.records} == {1}
    assert len(one.unmapped) < len(both.unmapped)
    assert parse_excel(LAYOUTS, sheets=["SPEC"]).by_key().keys() == one.by_key().keys()


def test_매핑되는_라벨이_값을_먼저_가져간다(layouts):
    """스키마에 없는 텍스트가 값을 채가면 진짜 라벨이 굶는다."""
    locs = {r.source_locator for r in layouts.records}
    assert "SPEC!E35" in locs                # Fail Position 의 값
    # 같은 셀을 두 라벨이 값으로 쓰지 않는다
    assert len(locs) == len({r.source_locator for r in layouts.records if r.found})


def test_xls_는_xlrd_경로로_보낸다(monkeypatch, tmp_path):
    """계약상 ParserType.EXCEL 은 xls 를 포함한다(src/contracts.py).
    openpyxl 은 .xls 를 못 읽으므로 확장자로 갈라 xlrd 경로를 타야 한다."""
    import openpyxl

    from src.parsers.text import excel as excel_mod

    made = openpyxl.Workbook()
    ws = made.active
    ws.title = "SPEC"
    ws["A1"], ws["B1"] = "Manufacturer", "FISHER"
    ws["A2"], ws["B2"] = "Model No.", "667-ED"

    called = []
    monkeypatch.setattr(excel_mod, "load_xls", lambda p: (called.append(p), made)[1])

    path = tmp_path / "old.xls"
    path.write_bytes(b"")                       # 내용은 load_xls 가 만든 것으로 대체된다
    by = excel_mod.parse_excel(str(path)).by_key()

    assert called == [str(path)]                # openpyxl 로 열지 않았다
    assert by["manufacturer"].raw_value == "FISHER"
    assert by["model_no"].raw_value == "667-ED"


@pytest.mark.skipif(__import__("importlib").util.find_spec("xlwt") is None,
                    reason="xlwt 미설치 — 구형 xls 를 만들 수 없다")
def test_구형_xls_왕복(tmp_path):
    """실제 .xls 왕복. xlwt 가 있을 때만 돈다.
    (설치 없이도 raw_file 의 실물 .xls 로 수동 확인함)"""
    import xlwt

    from src.parsers.text.xls_compat import load_xls

    path = tmp_path / "old.xls"
    wbx = xlwt.Workbook()
    sh = wbx.add_sheet("SPEC")
    sh.write(0, 0, "Manufacturer")
    sh.write(0, 1, "FISHER")
    wbx.save(str(path))

    assert load_xls(str(path)).sheetnames == ["SPEC"]
    assert parse_excel(str(path)).by_key()["manufacturer"].raw_value == "FISHER"


# ── 2단 양식에서 Max/Nor/Min 열이 새지 않는다 (실물 10FV079) ─────


def _twocol_sheet(path):
    """왼쪽 라벨·값 / 오른쪽 라벨·값 2단 양식 + 위쪽 서비스조건 블록.

    실물 `10FV079` 의 배치다. 서비스조건 머리글(NOR. 54열)이 시트 끝까지
    유효하면 아래쪽 행에서 오른쪽 단의 '라벨' 칸을 값으로 집는다.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.cell(1, 35).value, ws.cell(1, 44).value = "Units", "MIN."
    ws.cell(1, 54).value, ws.cell(1, 64).value = "NOR.", "MAX."

    ws.cell(2, 8).value = "Required Cv"                  # 블록 안 — Nor 열이 맞다
    ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=34)
    for col, v in [(35, "Cv"), (44, "1.16"), (54, "2.51"), (64, "12.2")]:
        ws.cell(2, col).value = v
    for a, b in [(44, 53), (54, 63), (64, 73)]:
        ws.merge_cells(start_row=2, start_column=a, end_row=2, end_column=b)

    ws.cell(9, 8).value = "Rated Cv"                     # 블록 밖 — 왼쪽 값이 맞다
    ws.merge_cells(start_row=9, start_column=8, end_row=9, end_column=22)
    ws.cell(9, 23).value = "26"
    ws.merge_cells(start_row=9, start_column=23, end_row=9, end_column=45)
    ws.cell(9, 53).value = "Body Color"                  # 오른쪽 단의 라벨
    ws.merge_cells(start_row=9, start_column=53, end_row=9, end_column=69)
    ws.cell(9, 70).value = "SILVER"
    wb.save(path)
    return path


def test_Nor_열은_블록_밖_행까지_따라가지_않는다(tmp_path):
    by = parse_excel(_twocol_sheet(str(tmp_path / "twocol.xlsx"))).by_key()
    assert by["rated_cv"].raw_value == "26"          # 'Body Color' 가 아니다
    assert by["required_cv"].raw_value == "2.51"     # 블록 안에서는 Nor 열이 이긴다


def test_가이드_부싱을_케이지_재질로_보지_않는다():
    """가이드 부싱은 케이지와 다른 부품이다.

    실물 10FV079 는 두 칸이 따로 있고 값도 다르며(`Guide Material` = SOLID
    STELLITE vs 케이지 316 SST), 10PV081 은 `Cage` 에만 값이 있다.
    킷은 갈린다 — d010·d011 은 Guide 행 값을 케이지 정답으로 적었다. 설계에 따라
    케이지가 가이드를 겸하기도 해서 도메인 판단이 필요하고, 그때까지는
    **미추출이 오답보다 낫다**는 원칙으로 빼 둔다.
    """
    ix = FieldIndex.load()
    for label in ("Guide Material", "Guide Bushing", "MATERIAL Guide Bushing"):
        hits = [h.key for h in ix.candidates(label)]
        assert "valve_cage_material" not in hits, (label, hits)
