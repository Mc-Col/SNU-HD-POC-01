# -*- coding: utf-8 -*-
"""라벨링 킷의 원문라벨 → 유사표현 사전 후보.

  python -m src.parsers.text.aliases_from_kit --kit ~/snu_ai/poc_team/labeling_kit.xlsx

킷 안내문:
  "원문라벨에 문서에 적혀 있던 항목명을 그대로 옮겨 적는다.
   이 값들이 모여 유사표현 사전이 된다."
  "괄호로 시작하면 표기 변종이 아니라 위치 설명으로 처리되어 사전에 들어가지 않는다."

결과는 readme/output_sample.xlsx 의 '유사 표현' 행에 붙여 넣고
`python tools/gen_schema.py` 로 재생성하면 반영된다. 이 도구는 스키마를 고치지 않는다.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from collections import OrderedDict
from dataclasses import dataclass, field as dc_field

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

import openpyxl                                                     # noqa: E402
import yaml                                                         # noqa: E402

from src.parsers.text.composite import CompositeIndex               # noqa: E402
from src.parsers.text.field_index import (                          # noqa: E402
    SCHEMA_PATH, FieldIndex, normalize_label,
)

SHEET = "라벨링"

# 라벨이 아니라 "문서에 항목이 없음"을 적어 둔 칸. 사전에 넣으면 안 된다.
_NO_LABEL = re.compile(r"^(N\s*/?\s*A|없음|해당없음)\b", re.I)
# 라벨 뒤에 붙은 설명 괄호. 붙여 쓴 괄호(Body Model(Type))는 라벨의 일부이므로 남긴다.
_TRAILING_NOTE = re.compile(r"\s+\([^)]*\)\s*$")
_HANGUL = re.compile(r"[가-힣]")


def clean_label(text: str) -> str:
    """원문라벨에서 설명 괄호를 떼어 낸다. 라벨이 아니면 빈 문자열."""
    t = " ".join(str(text or "").split())
    if not t or t.startswith("(") or _NO_LABEL.match(t):
        return ""
    m = _TRAILING_NOTE.search(t)
    if m and (_HANGUL.search(m.group(0)) or len(m.group(0)) > 14):
        t = t[:m.start()].strip()
    return t


@dataclass
class Candidate:
    text: str                       # 문서에 적힌 항목명 그대로
    docs: list[str] = dc_field(default_factory=list)

    @property
    def count(self) -> int:
        return len(self.docs)


@dataclass
class Result:
    by_field: "OrderedDict[str, list[Candidate]]" = dc_field(default_factory=OrderedDict)
    already: list[tuple[str, str]] = dc_field(default_factory=list)   # (필드, 표기)
    positional: list[tuple[str, str]] = dc_field(default_factory=list)  # 괄호 = 위치 설명
    unresolved_fields: list[str] = dc_field(default_factory=list)
    collisions: list[tuple[str, list[str]]] = dc_field(default_factory=list)

    @property
    def total(self) -> int:
        return sum(len(v) for v in self.by_field.values())


def _existing(path: str = SCHEMA_PATH) -> dict[str, set[str]]:
    """필드명 → 이미 등록된 표기(정규화). 표준명 자신도 포함한다."""
    with open(path, encoding="utf-8") as f:
        fields = yaml.safe_load(f)["fields"]
    out = {}
    for x in fields:
        known = {normalize_label(x["name"])}
        known |= {normalize_label(a) for a in x.get("aliases") or []}
        out[x["name"]] = known
    return out


def collect(kit_path: str, ix: FieldIndex | None = None) -> Result:
    ix = ix or FieldIndex.load()
    known = _existing()
    ws = openpyxl.load_workbook(kit_path, data_only=True)[SHEET]
    res = Result()

    cur = None
    for c in range(10, ws.max_column + 1):
        h1, h2 = ws.cell(1, c).value, ws.cell(2, c).value
        if h1:
            cur = str(h1).replace("※안전", "").strip().lstrip("· ").strip()
        if h2 != "원문라벨" or not cur:
            continue

        hit = ix.lookup(cur)
        if hit is None:
            if cur not in res.unresolved_fields:
                res.unresolved_fields.append(cur)
            continue
        seen = known.get(hit.name, set())
        bucket: "OrderedDict[str, Candidate]" = OrderedDict()

        for r in range(3, ws.max_row + 1):
            raw = ws.cell(r, c).value
            doc = str(ws.cell(r, 1).value or "")
            original = " ".join(str(raw or "").split())
            if not original:
                continue
            text = clean_label(original)
            if not text:                             # 위치 설명·항목 없음 표기
                res.positional.append((hit.name, original))
                continue
            n = normalize_label(text)
            if not n:
                continue
            if n in seen:
                res.already.append((hit.name, text))
                continue
            bucket.setdefault(n, Candidate(text)).docs.append(doc)

        if bucket:
            res.by_field.setdefault(hit.name, []).extend(
                sorted(bucket.values(), key=lambda x: (-x.count, x.text)))

    # 한 표기가 여러 필드에 걸리면 유사표현이 아니라 복합 라벨이다.
    # 예) "Size and Type" = 1" 1250R-GL → 바디 사이즈 + 모델번호
    cix = CompositeIndex.load()
    where: dict[str, list[str]] = {}
    for name, cands in res.by_field.items():
        for cd in cands:
            where.setdefault(normalize_label(cd.text), []).append(name)
    for name, cands in res.by_field.items():
        for cd in cands:
            n = normalize_label(cd.text)
            if len(where[n]) > 1 and cix.lookup(cd.text) is None:
                entry = (cd.text, sorted(set(where[n])))
                if entry not in res.collisions:
                    res.collisions.append(entry)
    return res


def render(res: Result) -> tuple[str, str]:
    L = ["# 유사표현 후보 — 라벨링 킷 원문라벨에서 수집", ""]
    L += [f"- 새 표기 **{res.total}건** · 필드 {len(res.by_field)}개",
          f"- 이미 등록됨 {len(res.already)}건 · 위치 설명이라 제외 {len(res.positional)}건", ""]
    if res.unresolved_fields:
        L += ["> 킷 이름이 스키마에 없어 반영 불가: "
              + ", ".join(f"`{x}`" for x in res.unresolved_fields), ""]

    if res.collisions:
        L += ["## ⚠ 유사표현으로 넣으면 안 되는 표기", "",
              "한 표기가 여러 필드에 걸린다. 그대로 넣으면 한 필드만 이기고 나머지는 굶는다.",
              "- 한 칸에 값이 여러 개면 → `schema/rules.yaml` 의 `composite_labels`",
              "- 구역(섹션)만 다른 같은 단어면 → 구역 인식이 필요하다 (아직 미구현)", "",
              "| 표기 | 걸린 필드 |", "|---|---|"]
        for t, names in res.collisions:
            L.append(f"| `{t}` | {' + '.join(names)} |")
        L += [""]

    L += ["## 필드별", "", "| 표준 필드 | 새 유사표현 | 건수 | 문서 |", "|---|---|---:|---|"]
    for name, cands in res.by_field.items():
        for cd in cands:
            L.append(f"| {name} | `{cd.text}` | {cd.count} | {', '.join(cd.docs)} |")

    if res.positional:
        L += ["", "## 위치 설명 (사전에 넣지 않음)", ""]
        for name, t in res.positional:
            L.append(f"- {name} — `{t}`")

    L += ["", "## 반영 방법", "",
          "1. `readme/output_sample.xlsx` 의 `유사 표현` 행에 필드별로 붙여 넣기",
          "2. `python tools/gen_schema.py` 로 `schema/fields.yaml` 재생성", ""]

    T = ["표준필드\t유사표현\t건수\t채택(O/X)"]
    for name, cands in res.by_field.items():
        for cd in cands:
            T.append(f"{name}\t{cd.text}\t{cd.count}\t")
    return "\n".join(L) + "\n", "\n".join(T) + "\n"


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="킷 원문라벨 → 유사표현 후보")
    ap.add_argument("--kit", required=True)
    ap.add_argument("--out", default="runs/alias_candidates.md")
    a = ap.parse_args(argv)

    res = collect(os.path.expanduser(a.kit))
    md, tsv = render(res)
    os.makedirs(os.path.dirname(os.path.abspath(a.out)) or ".", exist_ok=True)
    with open(a.out, "w", encoding="utf-8") as f:
        f.write(md)
    tsv_path = os.path.splitext(a.out)[0] + ".tsv"
    with open(tsv_path, "w", encoding="utf-8") as f:
        f.write(tsv)
    print(f"새 유사표현 {res.total}건 · 필드 {len(res.by_field)}개 "
          f"· 이미등록 {len(res.already)} · 위치설명 {len(res.positional)}")
    print(f"  {a.out}\n  {tsv_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
