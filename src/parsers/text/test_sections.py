# -*- coding: utf-8 -*-
"""③-a TEXT PARSER — 구역(section) 인식 검증.

무엇을 지키는 테스트인가
────────────────────────────────────────────────────────────────────
데이터시트는 부품별로 묶여 있고 같은 항목명이 묶음마다 되풀이된다.
구역 인식은 "이 라벨이 어느 묶음에 있는가" 를 읽어 **다른 부품의 값을
집지 않게** 막는 장치다. 여기서 지키는 것은 네 가지다.

  ① 이름표를 찾는다        엑셀=세로 병합 셀 · PDF=90도 회전 글자
  ② 범위를 옳게 나눈다      회전 이름표는 블록 중앙에 서고, 표 밖까지 늘어나면 안 된다
  ③ 2단 양식을 가른다       왼쪽 단 라벨은 왼쪽 이름표, 오른쪽 단은 오른쪽 이름표
  ④ 모르면 막지 않는다      구역이 없는 문서는 지금까지와 똑같이 동작한다

좌표·배치는 실물에서 그대로 옮겼다 (`52PV014` · `10FV079` · `44LV001`).
"""
import os
import sys

import pytest

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from src.parsers.text.field_index import FieldIndex                # noqa: E402
from src.parsers.text.sections import SectionIndex, SectionMap     # noqa: E402


# 사전은 실물 표기를 그대로 쓴다. rules.yaml 이 바뀌어도 이 테스트는 자기 사전으로 돈다.
DICT = SectionIndex(
    aliases={
        "GENERAL": "general",
        "SERVICE CONDITION": "service",
        "VALVE BODY / BONNET": "body",
        "TRIM": "trim",
        "ACTUATOR": "actuator",
        "POSITIONER": "positioner",
        "LIMIT SW": "none",
    },
    fields={
        "general": ["engineering_tag_no"],
        "service": ["normal_flow_rate"],
        "body": ["model_no", "valve_body_material"],
        "trim": ["valve_plug_material"],
        "actuator": ["actuator_fail_action"],
        "positioner": ["positioner_model_no"],
        "none": [],
    },
)


# ── 엑셀 — 세로 병합 셀이 이름표다 ─────────────────────────────


def _sheet():
    """10FV079 배치. 왼쪽 단(라벨 8열)과 오른쪽 단(라벨 53열)이 각자 이름표를 갖는다."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    for text, r0, r1, c0, c1 in [
        ("GENERAL", 7, 10, 1, 4),
        ("SERVICE CONDITION", 11, 23, 1, 4),
        ("VALVE BODY / BONNET", 24, 37, 1, 4),
        ("ACTUATOR", 49, 56, 1, 4),
        ("POSITIONER", 24, 30, 46, 49),      # 오른쪽 단 담당
        ("Authorized by", 40, 44, 1, 4),     # 사전에 없는 이름표 (정비보고서 서명란)
    ]:
        ws.cell(r0, c0).value = text
        ws.merge_cells(start_row=r0, start_column=c0, end_row=r1, end_column=c1)
    return ws


@pytest.fixture(scope="module")
def excel_map():
    return SectionMap.from_excel(_sheet(), DICT)


@pytest.mark.parametrize("row,col,want", [
    (24, 8, "body"),        # 밸브 본체 묶음의 라벨
    (50, 8, "actuator"),    # 액추에이터 묶음의 라벨 — 여기 'Model No.' 가 있다
    (12, 8, "service"),
    (26, 53, "positioner"),  # 오른쪽 단은 오른쪽 이름표에 붙는다
    (60, 8, None),          # 어느 이름표 범위에도 없다
    (24, 2, None),          # 이름표 자신보다 왼쪽 — 지배하지 않는다
])
def test_엑셀은_세로_병합_셀로_구역을_가른다(excel_map, row, col, want):
    assert excel_map.at(row, col) == want


def test_사전에_없는_이름표는_구역으로_보지_않는다(excel_map):
    """10PV018 정비보고서에는 'Authorized by' 같은 세로 병합 셀이 5개 있다.

    모양만 보고 구역으로 삼으면 그런 것까지 구역이 된다.
    """
    assert excel_map.at(42, 8) is None
    assert all(m.text != "Authorized by" for m in excel_map.marks)


# ── PDF — 90도 회전된 글자가 이름표다 ──────────────────────────


def _page(tmp_path):
    """52PV014 배치를 좌표까지 그대로 옮긴다.

    왼쪽 이름표 x=57 / 오른쪽 이름표 x=295, 라벨은 각각 x=90 · x=327.
    """
    import fitz
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)
    for x, y, text in [
        (66, 162, "GENERAL"),               # y=125~162 에 서는 세로 글자
        (66, 271, "SERVICE CONDITION"),
        (66, 417, "VALVE BODY / BONNET"),
        (66, 624, "ACTUATOR"),
        (303, 360, "POSITIONER"),           # 오른쪽 단 담당
        (303, 490, "LIMIT SW"),
    ]:
        page.insert_text((x, y), text, fontsize=7, rotate=90)
    path = str(tmp_path / "sections.pdf")
    doc.save(path)
    doc.close()
    return fitz.open(path)[0]


@pytest.fixture(scope="module")
def pdf_map(tmp_path_factory):
    return SectionMap.from_pdf(_page(tmp_path_factory.mktemp("sec")), DICT)


def test_PDF_는_회전된_글자를_이름표로_읽는다(pdf_map):
    assert {m.key for m in pdf_map.marks} == {
        "general", "service", "body", "actuator", "positioner", "none"}


@pytest.mark.parametrize("y,x,want", [
    (140, 90, "general"),
    (230, 90, "service"),
    (380, 90, "body"),
    (600, 90, "actuator"),
    (340, 327, "positioner"),   # 오른쪽 단 라벨은 오른쪽 이름표에 붙는다
    (470, 327, "none"),         # 우리 스키마 밖 묶음
])
def test_PDF_는_이웃_이름표의_중점으로_경계를_잡는다(pdf_map, y, x, want):
    """회전 이름표는 블록의 세로 '중앙' 에 선다.

    "이 이름표부터 다음 이름표까지" 로 자르면 한 블록씩 밀린다.
    """
    assert pdf_map.at(y, x) == want


def test_표_밖의_문서_머리글은_구역에_딸려가지_않는다(pdf_map):
    """실물 19FV077 의 `Valve Tag # : 19-FV-077` 은 y=29, 표보다 위에 있다.

    맨 위 이름표의 범위를 무한대로 늘리면 이런 머리글까지 그 구역이 되어
    태그가 통째로 버려진다 (실제로 3개 문서에서 태그를 잃었다).
    """
    assert pdf_map.at(29, 327) is None
    assert pdf_map.at(20, 90) is None


# ── 구역이 라벨 해석을 가른다 ──────────────────────────────────


@pytest.fixture(scope="module")
def ix():
    return FieldIndex.load()


def test_구역이_다른_부품의_항목을_버린다(ix):
    """액추에이터 묶음의 `Model No.`(880)를 밸브 모델로 집던 오답을 막는다."""
    assert ix.lookup("Model No.") is not None                       # 구역 모르면 종전대로
    assert ix.lookup("Model No.", DICT.allowed("actuator")) is None  # 액추에이터 묶음이면 버린다
    assert ix.lookup("Model No.", DICT.allowed("body")).key == "model_no"


def test_한_이름이_여러_필드에_걸리면_구역이_고른다(ix):
    """`Model` 은 밸브 본체 모델일 수도, 포지셔너 모델일 수도 있다."""
    assert ix.lookup("Model") is None                                # 애매하면 만들지 않는다
    assert ix.lookup("Model", DICT.allowed("body")).key == "model_no"
    assert ix.lookup("Model", DICT.allowed("positioner")).key == "positioner_model_no"


def test_표준명을_가진_필드가_우선한다(ix):
    """다른 필드가 이름을 빌려 써도, 구역을 모를 때는 이름 주인이 이긴다.

    이 규칙이 없으면 킷 열 이름(`MODEL NO.`)조차 애매해져 채점에서 빠진다.
    """
    hit = ix.lookup("MODEL NO.")
    assert hit is not None and hit.key == "model_no" and hit.matched_on == "name"


def test_스키마_밖_묶음에서는_아무_값도_만들지_않는다(ix):
    """LIMIT SW · ACCESSORIES 처럼 우리 28필드에 대응이 없는 묶음."""
    assert DICT.allowed("none") == set()
    assert ix.lookup("Model No.", DICT.allowed("none")) is None
