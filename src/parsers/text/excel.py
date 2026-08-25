# -*- coding: utf-8 -*-
"""③-a TEXT PARSER (엑셀) — 헤더 텍스트를 표준 컬럼에 붙인다.

셀 좌표로 매핑하지 않는다. 벤더 양식은 디테일이 바뀌므로 헤더 텍스트가 기준이다.
값은 라벨의 오른쪽 → 아래 순서로 찾는다.
"""
from __future__ import annotations

from dataclasses import dataclass, field as dc_field

import openpyxl
from openpyxl.utils import get_column_letter

from src.contracts import ParserType, RawExtraction

from .columns import anchor_from
from .composite import CompositeIndex, try_split
from .field_index import FieldIndex
from .units import UnitIndex
from .xls_compat import load_xls

SCAN_RIGHT = 4          # 라벨 오른쪽으로 몇 칸까지 값을 찾는가
SCAN_DOWN = 2           # 오른쪽에 없으면 아래로 몇 칸까지
RIGHT_PROBE = 15        # "이 셀이 라벨인가" 를 볼 때만 쓰는 넓은 탐침.
                        # 값 탐색(SCAN_RIGHT)을 넓히면 노이즈가 늘지만, 라벨 판별은
                        # 넓게 봐야 한다 — 44LV001 은 라벨 B열·값 S열로 11칸 떨어져 있다
MAX_LABEL_LEN = 60      # 이보다 긴 문자열은 라벨이 아니라 본문으로 본다


@dataclass
class UnmappedLabel:
    """표준 컬럼에 붙지 못한 라벨. 유사표현 사전을 키우는 재료가 된다."""
    text: str
    source_locator: str
    neighbor_value: str


@dataclass
class TextParseResult:
    records: list[RawExtraction] = dc_field(default_factory=list)
    unmapped: list[UnmappedLabel] = dc_field(default_factory=list)

    def by_key(self) -> dict[str, RawExtraction]:
        return {r.field_key: r for r in self.records}


def _merged_index(ws) -> dict[tuple[int, int], tuple[tuple[int, int], object]]:
    """병합 셀의 모든 좌표 → (앵커 좌표, 병합범위)."""
    out: dict[tuple[int, int], tuple[tuple[int, int], object]] = {}
    for rng in ws.merged_cells.ranges:
        anchor = (rng.min_row, rng.min_col)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                out[(r, c)] = (anchor, rng)
    return out


def _text(v: object) -> str:
    return "" if v is None else str(v).strip()


def parse_excel(
    path: str,
    index: FieldIndex | None = None,
    composite: CompositeIndex | None = None,
    sheets: list[str | int] | None = None,
    units: UnitIndex | None = None,
) -> TextParseResult:
    """엑셀 파일 하나 → RawExtraction[] + 미매핑 라벨.

    sheets 는 시트 이름 또는 1-based 순번. None 이면 전체.
    Triage 가 사양표 시트를 지정하면 그것만 본다 (사진·이력 시트 노이즈 배제).
    """
    ix = index or FieldIndex.load()
    cix = composite if composite is not None else CompositeIndex.load()
    uix = units if units is not None else UnitIndex.load()
    wb = (load_xls(path) if path.lower().endswith(".xls")
          else openpyxl.load_workbook(path, data_only=True))
    result = TextParseResult()
    seen: set[str] = set()                      # 먼저 찾은 값이 이긴다

    for page, ws in enumerate(wb.worksheets, start=1):
        if sheets is not None and page not in sheets and ws.title not in sheets:
            continue
        merged = _merged_index(ws)
        consumed: set[tuple[int, int]] = set()   # 값으로 이미 쓰인 셀은 라벨이 아니다

        def cell_value(r: int, c: int) -> object:
            hit = merged.get((r, c))
            if hit is not None:
                ar, ac = hit[0]
                return ws.cell(ar, ac).value
            return ws.cell(r, c).value

        def locator(r: int, c: int) -> str:
            return f"{ws.title}!{get_column_letter(c)}{r}"

        # 공정 조건 Max/Nor/Min 열 머리글 → 그 아래 행에만 Normal 열을 적용한다.
        # 시트 전체에 걸면 무관한 행이 엉뚱한 열의 값을 집는다.
        nor_by_row: dict[int, int] = {}
        cur: int | None = None
        for row in ws.iter_rows():
            if not row:
                continue
            found = anchor_from([(c.value, c.column) for c in row if c.value is not None])
            if found is not None:
                cur = int(found)
                continue                     # 머리글 행 자체는 값이 없다
            if cur is not None:
                nor_by_row[row[0].row] = cur

        # 1패스: 표준 컬럼에 붙는 라벨. 2패스: 나머지 라벨 후보(미매핑 수집).
        # 매핑되는 라벨이 값을 먼저 claim 해야 엉뚱한 텍스트가 값을 채가지 않는다.
        candidates = [c for row in ws.iter_rows() for c in row]
        known = [c for c in candidates
                 if isinstance(c.value, str)
                 and (ix.lookup(_text(c.value)) is not None
                      or cix.lookup(_text(c.value)) is not None)]
        rest = [c for c in candidates if c not in known]

        for cell in known + rest:
                r, c = cell.row, cell.column
                if (r, c) in consumed:
                    continue
                # 병합 영역이면 앵커에서만 처리한다 (같은 라벨 중복 방지)
                span = merged.get((r, c))
                if span is not None and span[0] != (r, c):
                    continue
                if not isinstance(cell.value, str):      # 숫자·날짜는 라벨이 아니다
                    continue

                label = _text(cell.value)
                if not label or len(label) > MAX_LABEL_LEN:
                    continue
                if not any(ch.isalpha() for ch in label):
                    continue

                hit = ix.lookup(label)
                value, vpos = _find_value(cell_value, ix, merged, r, c, span, consumed,
                                          uix, nor_by_row.get(r))
                if value:
                    consumed.add(vpos)

                # 라벨 하나가 여러 필드를 덮는 칸이면 쪼갠다
                split = try_split(cix, ix, label, value)
                if split is not None:
                    ok, pending = split
                    for pc in ok:
                        if pc.field_key in seen:
                            continue
                        seen.add(pc.field_key)
                        result.records.append(RawExtraction(
                            field_key=pc.field_key,
                            raw_value=pc.value,
                            raw_label=pc.label,
                            page=page,
                            confidence=pc.confidence,
                            parser=ParserType.EXCEL,
                            source_locator=locator(*vpos),
                            note=f"복합 라벨 '{label}' 분해",
                        ))
                    for pc in pending:
                        result.unmapped.append(UnmappedLabel(
                            pc.label, locator(*vpos), pc.value))
                    continue

                if hit is None:
                    if value:
                        result.unmapped.append(
                            UnmappedLabel(label, locator(r, c), value)
                        )
                    continue

                if hit.key in seen:
                    continue
                seen.add(hit.key)

                found = bool(value)
                result.records.append(RawExtraction(
                    field_key=hit.key,
                    raw_value=value if found else None,
                    raw_label=label,
                    page=page,
                    confidence=hit.confidence if found else 0.0,
                    parser=ParserType.EXCEL,
                    source_locator=locator(*vpos) if found else locator(r, c),
                    note="" if found else "라벨은 찾았으나 값 셀이 비어 있음",
                ))

    result.records.sort(key=lambda x: x.field_key)
    result.unmapped.sort(key=lambda x: (x.source_locator, x.text))
    return result


def _find_value(cell_value, ix: FieldIndex, merged, r: int, c: int, span,
                consumed: set[tuple[int, int]] | None = None,
                units: UnitIndex | None = None,
                nor_col: int | None = None):
    """라벨 오른쪽 → 아래 순서로 값 셀을 찾는다. 다른 라벨을 만나면 멈춘다.

    Max/Nor/Min 열을 아는 행이면 Normal 열을 먼저 본다 (탐색 폭 밖이어도).
    """
    taken = consumed or set()
    uix = units
    rng = span[1] if span else None
    right_from = rng.max_col if rng else c
    down_from = rng.max_row if rng else r

    if nor_col is not None and nor_col > right_from and _nor_applies(merged, r, nor_col):
        pos = (r, nor_col)
        v = _text(cell_value(*pos))
        if v and pos not in taken and not (uix and uix.is_unit(v)) \
                and ix.lookup(v) is None:
            return v, pos

    for dc in range(1, SCAN_RIGHT + 1):
        pos = (r, right_from + dc)
        if pos in taken:
            continue
        v = _text(cell_value(*pos))
        if not v:
            continue
        if uix is not None and uix.is_unit(v):  # 단위 칸은 건너뛴다
            continue
        if ix.lookup(v) is not None:            # 값이 아니라 다음 라벨이다
            break
        return v, pos

    for dr in range(1, SCAN_DOWN + 1):
        pos = (down_from + dr, c)
        if pos in taken:
            continue
        v = _text(cell_value(*pos))
        if not v:
            continue
        if uix is not None and uix.is_unit(v):
            continue
        if ix.lookup(v) is not None:
            break
        if _has_own_value(cell_value, pos, uix, nor_col):
            break            # 자기 값을 오른쪽에 달고 있으면 그건 다음 행의 라벨이다
        return v, pos

    return "", (r, c)


def _nor_applies(merged, r: int, nor_col: int) -> bool:
    """이 행이 Max/Nor/Min 블록 안에 있는가.

    머리글은 한 번 나오면 그 아래 모든 행에 붙어 버린다. 2단 양식(왼쪽 라벨·값,
    오른쪽 라벨·값)에서는 Nor 열이 아래로 내려가다 오른쪽 블록의 라벨 칸을
    관통하고, 그 라벨 텍스트가 값으로 잡힌다 — 실물 `10FV079` 에서 4건.

    판별자는 칸 경계다. 블록 안의 행은 Nor 열에서 칸이 시작하지만,
    블록 밖의 행은 다른 열에서 시작한 병합칸이 그 자리를 지나갈 뿐이다.
    """
    span = merged.get((r, nor_col))
    return span is None or span[0][1] == nor_col


def _has_own_value(cell_value, pos: tuple[int, int], uix,
                   nor_col: int | None = None) -> bool:
    """이 셀이 오른쪽에 자기 값을 달고 있는가 (= 다음 행의 라벨이다)."""
    r, c = pos
    cols = list(range(c + 1, c + 1 + RIGHT_PROBE))
    if nor_col is not None and nor_col > c:
        cols.append(nor_col)
    for cc in cols:
        v = _text(cell_value(r, cc))
        if v and not (uix and uix.is_unit(v)):
            return True
    return False
