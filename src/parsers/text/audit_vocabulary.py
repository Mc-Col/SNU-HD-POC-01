# -*- coding: utf-8 -*-
"""허용 어휘 감사 — 정답지 값이 표준값에 닿는가.

  python -m src.parsers.text.audit_vocabulary --kit readme/labeling_kit.xlsx

무엇을 재는가
  `schema/rules.yaml` 의 `enum_allowed_values` 는 두 갈래다.

      correct     value_aliases 의 `to` 값이 곧 허용값. 값을 표준형으로 바꾼다
      flag_only   값을 바꾸지 않고 어휘 밖이면 확인필요로 표시만 한다 (재질·안전)

  **정답지 값이 어느 쪽에도 닿지 못하면** 그 칸은 판독이 완벽해도 정확으로
  채점되지 않고, 화면에서는 맞는 값이 매번 확인필요로 뜬다.

왜 이 감사가 필요한가
  `flag_only` 는 값을 건드리지 않으므로 **목록을 늘리는 데 위험이 0**이다.
  반대로 빠져 있으면 손해가 생긴다 — 검토자가 맞는 값을 매번 확인하게 되고,
  그것이 반복되면 진짜 경고도 흘려보게 된다. 목록은 사람이 문서에서 읽은
  값으로 자란다(Loop C). 이 도구가 그 자랄 자리를 찾아 준다.

  ⚠ **감사는 값을 고치라고 말하지 않는다.** `316 SS` 와 `316L SST` 는 다른
    강종이고 `CS`·`C5` 는 한 글자 차이로 탄소강과 크롬몰리강이다. 이 도구는
    "목록에 없다" 만 보고하고, 넣을지는 사람이 판단한다.

  ⚠ **모든 미달을 메우는 것이 목표가 아니다.** 근거 없는 값을 만드는 별칭은
    넣지 않는다 — `ACTUATOR TYPE` 의 `STD`(양식이 "표준품" 이라고만 적은 것)를
    `DIAPHRAGM` 으로 바꾸는 것이 그 예다(철학 4).

2026-08-26 첫 실행 — 정답 14칸이 어휘 밖이었고 13칸을 목록에 넣어 1칸이 남았다.
"""
from __future__ import annotations

import argparse
import collections
import os
import sys

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

import openpyxl                                                    # noqa: E402
import yaml                                                        # noqa: E402

from src import schema                                             # noqa: E402
from src.parsers.text.field_index import FieldIndex                # noqa: E402
from src.parsers.text.sections import SectionIndex                 # noqa: E402

SHEET = "라벨링"
SKIP = {"N/A", "NA", "판독불가", ""}
RULES = os.path.join(ROOT, "schema", "rules.yaml")


def _enum() -> tuple[set[str], dict[str, set[str]]]:
    doc = yaml.safe_load(open(RULES, encoding="utf-8")) or {}
    e = doc.get("enum_allowed_values") or {}
    correct = set(e.get("correct") or [])
    flag = {k: set(v) for k, v in (e.get("flag_only") or {}).items()
            if isinstance(v, list)}
    return correct, flag


def allowed(field_key: str, flag: dict[str, set[str]]) -> set[str]:
    """이 필드의 표준값 집합.

    `correct` 필드의 허용값은 **value_aliases 의 `to` 값이 곧 그것이다** —
    한 곳에만 적어 두 목록이 어긋나지 않게 한다(rules.yaml 주석).
    `compare_only` 별칭은 값을 바꾸지 않는 대조용이므로 제외한다.
    """
    tos = {str(m["to"]) for m in schema.value_aliases(field_key)
           if m.get("to") and not m.get("compare_only")}
    return tos | flag.get(field_key, set())


def reaches(field_key: str, value: str, std: set[str]) -> bool:
    """정답 값이 표준값이거나, 별칭을 거쳐 표준값에 닿는가.

    대조는 `norm_alias` 로 한다 — 필드에 따라 구두점을 살린다(치수).
    """
    probe = schema.norm_alias(value, field_key)
    if probe in {schema.norm_alias(x, field_key) for x in std}:
        return True
    for m in schema.value_aliases(field_key):
        if m.get("compare_only"):
            continue
        if probe in {schema.norm_alias(c, field_key) for c in m.get("from", [])}:
            return True
    return False


def audit(kit_path: str) -> dict[str, collections.Counter]:
    """필드별로 어휘 밖 정답 값을 센다. 값이 없는 칸(N/A)은 세지 않는다."""
    correct, flag = _enum()
    six = SectionIndex.load()
    ix = FieldIndex.load(section_names=six.name_map())
    ws = openpyxl.load_workbook(kit_path, data_only=True)[SHEET]

    cols, cur = [], None
    for c in range(10, ws.max_column + 1):
        h1, h2 = ws.cell(1, c).value, ws.cell(2, c).value
        if h1:
            cur = str(h1).replace("※안전", "").strip().lstrip("· ").strip()
        if h2 != "정답값" or not cur:
            continue
        hit = ix.lookup(cur)
        if hit:
            cols.append((c, hit.key))

    out: dict[str, collections.Counter] = collections.defaultdict(collections.Counter)
    for r in range(3, ws.max_row + 1):
        if not ws.cell(r, 2).value:
            continue
        for c, key in cols:
            if key not in correct and key not in flag:
                continue                      # 어휘 규칙이 없는 필드는 대상이 아니다
            v = ws.cell(r, c).value
            # 엑셀 왕복으로 숫자가 된 값을 되살린다 (195 → 195.0)
            t = (str(int(v)) if isinstance(v, float) and v.is_integer()
                 else str(v or "")).strip()
            if t.upper() in SKIP:
                continue
            std = allowed(key, flag)
            if not std or reaches(key, t, std):
                continue
            out[key][t] += 1
    return dict(out)


def render(gaps: dict[str, collections.Counter]) -> str:
    total = sum(sum(c.values()) for c in gaps.values())
    L = ["# 허용 어휘 감사 (정답지 대비)", "",
         f"- 어휘 밖 정답 **{total}칸** · {len(gaps)}필드", ""]
    if not gaps:
        L += ["정답지 값이 전부 표준값에 닿는다.", ""]
        return "\n".join(L)
    L += ["> 넣을지는 사람이 판단한다. 한 글자 차이가 다른 재질인 필드가 있다.",
          "", "| 필드 | 값 | 건수 |", "|---|---|---|"]
    for key, cnt in sorted(gaps.items(), key=lambda x: -sum(x[1].values())):
        for v, n in cnt.most_common():
            L.append(f"| {schema.get(key).name} | `{v}` | {n} |")
    return "\n".join(L) + "\n"


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="허용 어휘 감사 — 정답지 대비")
    ap.add_argument("--kit", required=True)
    ap.add_argument("--out", default="runs/vocab_audit.md")
    a = ap.parse_args(argv)

    gaps = audit(a.kit)
    md = render(gaps)
    os.makedirs(os.path.dirname(a.out) or ".", exist_ok=True)
    with open(a.out, "w", encoding="utf-8") as fh:
        fh.write(md)
    total = sum(sum(c.values()) for c in gaps.values())
    print(f"어휘 밖 정답 {total}칸 · {len(gaps)}필드")
    print(f"  {a.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
