# -*- coding: utf-8 -*-
"""구형 .xls 를 openpyxl 워크북으로 읽어 들인다.

계약상 ParserType.EXCEL 은 xlsx / xlsm / xls 를 포함한다(src/contracts.py).
openpyxl 은 .xls 를 못 읽으므로 xlrd 로 읽어 메모리에서 옮겨 담는다.
값과 병합 정보만 옮긴다 — 파서가 쓰는 건 그 둘뿐이다.
"""
from __future__ import annotations

import openpyxl
import xlrd

MAX_SHEET_NAME = 31


def load_xls(path: str) -> openpyxl.Workbook:
    try:
        book = xlrd.open_workbook(path, formatting_info=True)
        merged_ok = True
    except NotImplementedError:
        book = xlrd.open_workbook(path)      # 일부 파일은 서식 정보를 못 읽는다
        merged_ok = False

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sh in book.sheets():
        ws = wb.create_sheet(title=str(sh.name)[:MAX_SHEET_NAME])
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                v = sh.cell_value(r, c)
                if v == "" or v is None:
                    continue
                ws.cell(r + 1, c + 1).value = v
        if merged_ok:
            for rlo, rhi, clo, chi in sh.merged_cells:   # hi 는 배타적
                if rhi - rlo > 0 and chi - clo > 0:
                    ws.merge_cells(start_row=rlo + 1, start_column=clo + 1,
                                   end_row=rhi, end_column=chi)
    if not wb.worksheets:
        wb.create_sheet(title="empty")
    return wb
