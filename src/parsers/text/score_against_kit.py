# -*- coding: utf-8 -*-
"""골든셋(라벨링 킷) 대비 텍스트 파서 채점 — 내 모듈 자기 검증.

  python -m src.parsers.text.score_against_kit \
      --kit ~/snu_ai/poc_team/labeling_kit.xlsx --root "raw_file/서경빈 선임님"

이 도구는 src/parsers/text 만 잰다. 파이프라인 전체 평가 하네스(eval/)와 다르다.
정답이 아직 AI 초안이면 숫자는 잠정치다 — 사람 검수 뒤 다시 돌릴 것.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass, field as dc_field

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

import openpyxl                                                    # noqa: E402

from src.parsers.text.composite import CompositeIndex              # noqa: E402
from src.parsers.text.excel import parse_excel                     # noqa: E402
from src.parsers.text.field_index import FieldIndex, normalize_label  # noqa: E402
from src.parsers.text.sections import SectionIndex                 # noqa: E402
from src.parsers.text.pdf_text import parse_pdf_text               # noqa: E402
from src.parsers.text.units import UnitIndex                       # noqa: E402

# 값 대조는 팀의 평가 하네스 규칙을 그대로 쓴다 (단위 표기·로마자·대소문자).
# 여기서 다시 만들면 채점기와 하네스가 서로 다른 답을 내놓는다.
from eval.compare import same as _same                             # noqa: E402
from src import schema as _schema                                  # noqa: E402
from src.parsers.text.crosscheck import numeric_flag as _numeric   # noqa: E402


def _standardize(field_key: str, value: object) -> str:
    """schema/rules.yaml 의 표기 매핑을 적용한 값.

    킷은 FAIL ACTION 만 표준값으로 적게 되어 있다(킷 기입 안내). 파서는 문서
    원문(`VALVE CLOSE`)을 내므로, 사전을 적용하지 않으면 영원히 오답으로 잡힌다.
    """
    raw = str(value or "").strip()
    probe = _schema.norm_label(raw)
    for m in _schema.value_aliases(field_key):
        if probe in {_schema.norm_label(c) for c in m.get("from", [])}:
            return str(m["to"])
    return raw


# ── ④ Normalize 를 별도 축으로 함께 잰다 ─────────────────────────────
#
# 왜 기본 판정에 쓰지 않는가
#   규칙이 값을 고쳐 주면 **파서 결함이 규칙에 가려진다.** `OPEN` 을 집었든
#   `Body Color` 를 집었든 규칙을 통과한 뒤에는 구별이 안 될 수 있다. 그래서
#   파서 판정(정확/표기차이/정규화대기/오답)은 손대지 않고, 규칙이 메워 주는지를
#   **규칙해소 / 규칙공백** 이라는 다른 축으로 센다.
#
# 왜 필요한가
#   그 축이 없으면 정규화대기 칸이 전부 "할 일" 처럼 보인다. 2026-08-26 실측에서
#   정규화대기 52칸은 **전부 이미 규칙이 처리하는 것**이었다 — 할 일이 아니라
#   이미 끝난 일이다. 남는 것만 골라내야 다음 작업을 정할 수 있다.

_UNSET = object()
_NORM: object = _UNSET


def _normalizer():
    """④ Normalize 를 지연 로딩한다.

    `src/pipeline.py` 는 내 소유가 아니므로 **읽어서 쓰기만** 한다 —
    `eval/harness.py` · `tools/check_kit.py` 와 같은 방식이다. 지연 로딩인
    이유는 파이프라인이 파서·검증기를 다 끌고 오기 때문이다. 채점만 할 때
    그 비용을 물지 않는다.

    못 가져오면 None 을 돌리고 **사유를 남긴다**(철학 5). 채점 자체는 계속된다.
    """
    global _NORM
    if _NORM is _UNSET:
        try:
            from src.pipeline import DefaultNormalize
            _NORM = DefaultNormalize()
        except Exception as e:                                  # pragma: no cover
            _NORM = None
            print(f"[주의] ④ Normalize 를 불러오지 못했다 — "
                  f"규칙해소 집계를 생략한다: {type(e).__name__}: {e}", file=sys.stderr)
    return _NORM


def _normalized(field_key: str, ex: object) -> str:
    """파서가 낸 것을 ④ Normalize 에 통과시킨 값.

    ⚠ `RawExtraction` 을 통째로 넘긴다. 값만 넘기면 안 되는 규칙이 있다 —
    FAIL 어간 판정은 **라벨**을 본다(`Air Fails Valve to : Close` → `FAIL CLOSE`).
    값 `CLOSE` 만 주면 방향을 못 읽어 그대로 나온다.
    """
    n, f = _normalizer(), _schema.get(field_key)
    if n is None or f is None or ex is None:
        return ""
    try:
        value, _trace = n.run(ex, f)
    except Exception:
        return ""                       # 규칙이 죽어도 채점은 계속한다
    return str(value or "").strip()


SHEET = "라벨링"
SKIP_VALUES = {"N/A", "NA", "판독불가", ""}      # 정답이 없는 칸 — 채점 제외
EXCEL_EXT = {".xlsx", ".xlsm"}


@dataclass
class Cell:
    doc_id: str
    field_key: str
    field_name: str
    truth: str
    uncertain: bool              # 정답 앞에 '?' 가 붙어 있었다
    got: str | None = None
    verdict: str = "미추출"       # 정확 / 표기차이 / 정규화대기 / 오답 / 미추출
    labeler: str = ""            # 이 정답을 만든 사람 — 집계를 나누기 위해
    normalized: str = ""         # ④ Normalize 통과값 (판정에 쓰지 않는다)
    rule_state: str = ""         # 규칙해소 / 규칙공백 — 파서 판정과 별도 축


@dataclass
class Score:
    cells: list[Cell] = dc_field(default_factory=list)
    docs: list[tuple[str, str, str]] = dc_field(default_factory=list)   # id, 파일, 상태
    unscorable_fields: list[str] = dc_field(default_factory=list)

    def counts(self, labeler: str | None = None) -> dict[str, int]:
        """판정별 개수. labeler 를 주면 그 사람이 만든 정답만 센다."""
        out = {"정확": 0, "표기차이": 0, "정규화대기": 0, "오답": 0, "미추출": 0}
        for c in self.cells:
            if labeler is not None and c.labeler != labeler:
                continue
            out[c.verdict] += 1
        return out

    def rule_counts(self, labeler: str | None = None) -> dict[str, int]:
        """규칙 축 개수. 정규화대기·오답 칸만 대상이다(그 밖은 빈 문자열)."""
        out = {"규칙해소": 0, "규칙공백": 0}
        for c in self.cells:
            if labeler is not None and c.labeler != labeler:
                continue
            if c.rule_state in out:
                out[c.rule_state] += 1
        return out

    def gaps(self) -> list["Cell"]:
        """규칙공백 칸 — 다음에 실제로 손봐야 할 것만 남는다."""
        return [c for c in self.cells if c.rule_state == "규칙공백"]

    def labelers(self) -> list[str]:
        """등장 순서대로의 라벨러 목록 (같은 입력 → 같은 출력)."""
        seen: list[str] = []
        for c in self.cells:
            if c.labeler not in seen:
                seen.append(c.labeler)
        return seen


def _cell_text(v: object) -> str:
    """킷 셀 값을 문자로. 엑셀 왕복으로 숫자가 된 값을 되살린다.

    사람이 킷을 엑셀에서 열어 저장하면 텍스트 "195" 가 숫자 195 로 바뀌고
    openpyxl 은 195.0 으로 읽는다. 그대로 대조하면 `195.0` vs 문서의
    `195 (Cg=4040 → 7580)` 이 달라져 맞는 값이 오답으로 잡힌다
    (2026-08-25 실제로 발생). 정수인 실수는 소수점을 떼고 본다.
    """
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v or "").strip()


def _norm(v: object) -> str:
    return re.sub(r"\s+", " ", str(v or "").strip()).upper()


def _loose(v: object) -> str:
    return re.sub(r"[^0-9A-Z]", "", str(v or "").upper())


def _contains(outer: object, inner: object) -> bool:
    """한쪽이 다른 쪽을 통째로 담고 있는가 (공백·기호 무시)."""
    a, b = _loose(outer), _loose(inner)
    return bool(a) and bool(b) and b in a


def read_kit(path: str, ix: FieldIndex) -> tuple[list[dict], list[tuple[int, str]], list[str]]:
    """킷 → (문서 행 목록, 채점 가능한 (열, field_key) 목록, 대응 미확정 필드명)"""
    ws = openpyxl.load_workbook(path, data_only=True)[SHEET]

    cols, unresolved, cur = [], [], None
    for c in range(10, ws.max_column + 1):
        h1, h2 = ws.cell(1, c).value, ws.cell(2, c).value
        if h1:
            cur = str(h1).replace("※안전", "").strip().lstrip("· ").strip()
        if h2 != "정답값" or not cur:
            continue
        hit = ix.lookup(cur)
        if hit is None:
            unresolved.append(cur)          # 킷 이름이 스키마에 없다 (예: RATED CV)
        else:
            cols.append((c, hit.key, hit.name))

    rows = []
    for r in range(3, ws.max_row + 1):
        fn = ws.cell(r, 2).value
        if not fn:
            continue
        rows.append({
            "row": r,
            "doc_id": str(ws.cell(r, 1).value or ""),
            "file": str(fn).strip(),
            "fmt": str(ws.cell(r, 3).value or ""),
            "labeler": str(ws.cell(r, 8).value or "").strip() or "미기재",
            "cls": str(ws.cell(r, 5).value or ""),
            "spec_page": ws.cell(r, 7).value,
            "truth": {key: (ws.cell(r, c).value, name) for c, key, name in cols},
        })
    return rows, cols, unresolved


def find_file(root: str, name: str) -> str | None:
    for dirpath, _, files in os.walk(root):
        if name in files:
            return os.path.join(dirpath, name)
    return None


def score(kit_path: str, root: str) -> Score:
    six = SectionIndex.load()
    # 구역 접두어를 뗀 표기까지 쓰는 인덱스로 채점한다 — 파서가 실제로 쓰는 것과 같게
    ix, cix = FieldIndex.load(section_names=six.name_map()), CompositeIndex.load()
    rows, _, unresolved = read_kit(kit_path, ix)
    sc = Score(unscorable_fields=unresolved)

    for row in rows:
        path = find_file(root, row["file"])
        if row["cls"] == "out_of_scope":
            sc.docs.append((row["doc_id"], row["file"], "제외(out_of_scope)"))
            continue
        if path is None:
            sc.docs.append((row["doc_id"], row["file"], "파일 없음"))
            continue

        ext = os.path.splitext(path)[1].lower()
        page = row["spec_page"]
        try:
            if ext in EXCEL_EXT:
                res = parse_excel(path, index=ix, composite=cix,
                                  sheets=[int(page)] if page else None)
            elif ext == ".pdf":
                res = parse_pdf_text(path, index=ix, composite=cix,
                                     pages=[int(page)] if page else None)
            else:
                sc.docs.append((row["doc_id"], row["file"], f"{ext} 미지원 — VLM 담당"))
                continue
        except Exception as e:
            sc.docs.append((row["doc_id"], row["file"], f"파싱 실패: {e}"))
            continue

        got = {r.field_key: r.raw_value for r in res.records if r.found}
        # 규칙해소 축은 RawExtraction 통째로 필요하다 (라벨을 보는 규칙이 있다)
        raws = {r.field_key: r for r in res.records if r.found}
        sc.docs.append((row["doc_id"], row["file"], "채점"))

        for key, (truth, name) in row["truth"].items():
            t = _cell_text(truth)
            uncertain = t.startswith("?")
            if uncertain:
                t = t[1:].strip()
            if _norm(t) in SKIP_VALUES:
                continue
            g = got.get(key)
            cell = Cell(row["doc_id"], labeler=row["labeler"], field_key=key,
                        field_name=name, truth=t, uncertain=uncertain, got=g)
            if g is None:
                cell.verdict = "미추출"
            elif _norm(g) == _norm(t):
                cell.verdict = "정확"
            elif _loose(g) == _loose(t) or _same(t, g, _numeric(key)):
                # 단위 표기·로마자·대소문자 차이는 감점하지 않는다 (eval/compare 규칙).
                #   예) 정답 "160 ℃" vs 파서 "160.0" → 같은 값이다
                cell.verdict = "표기차이"
            elif (_same(_standardize(key, t), _standardize(key, g), _numeric(key))
                  or _contains(g, t) or _contains(t, g)):
                # 표기 매핑 사전(rules.yaml)을 거쳐야 같아지는 것도 여기다.
                # **양쪽 모두** 접어서 비교한다 — 정답지도 표기가 갈리기 때문이다
                # (`600#` · `ANSI CLASS 300` · `300` 이 정답 안에 섞여 있다).
                # 파서는 문서 원문을 낸다. 표준값으로 바꾸는 것은 ④ Normalize 몫.
                #   문서가 더 짧을 수도(원문 "OPEN" → 표준 "FAIL OPEN"),
                #   더 길 수도 있다(원문 "195 (Cg=4040)" → 표준 "195").
                cell.verdict = "정규화대기"
            else:
                cell.verdict = "오답"

            # ── 별도 축 — 이 칸을 규칙이 메워 주는가 ──────────────────
            # 파서가 맞는 칸을 집었어도 표준값이 아니면 마스터에 그대로 못 넣는다.
            # 그것이 규칙으로 해결되는지, 규칙이 없는지를 여기서 가른다.
            # 이미 맞은 칸(정확·표기차이)과 아예 못 집은 칸(미추출)은 대상이 아니다.
            if cell.verdict in ("정규화대기", "오답"):
                nv = _normalized(key, raws.get(key))
                cell.normalized = nv
                hit = bool(nv) and (_norm(nv) == _norm(t)
                                    or _loose(nv) == _loose(t)
                                    or _same(t, nv, _numeric(key)))
                cell.rule_state = "규칙해소" if hit else "규칙공백"

            sc.cells.append(cell)
    return sc


def render(sc: Score) -> str:
    n = sc.counts()
    total = sum(n.values())
    L = ["# 텍스트 파서 채점 (골든셋 대비)", ""]
    L += [f"- 채점 칸 **{total}개** · 정확 {n['정확']} · 표기차이 {n['표기차이']} "
          f"· 정규화대기 {n['정규화대기']} · 오답 {n['오답']} · 미추출 {n['미추출']}"]
    if total:
        grabbed = n["정확"] + n["표기차이"] + n["정규화대기"]
        L += [f"- **파서 관점 성공률 {grabbed / total * 100:.0f}%** — 맞는 칸을 집었는가. "
              f"파서의 책임 범위는 여기까지다",
              f"- 완전 일치율 {(n['정확'] + n['표기차이']) / total * 100:.0f}% — "
              f"파서 원문 그대로. 표준값 변환 전이다"]
        r = sc.rule_counts()
        after = n["정확"] + n["표기차이"] + r["규칙해소"]
        L += [f"- **정규화 후 일치율 {after / total * 100:.0f}%** — ④ Normalize 를 "
              f"통과시킨 값 기준. 규칙해소 {r['규칙해소']}칸 · **규칙공백 "
              f"{r['규칙공백']}칸**",
              "",
              "> 숫자를 셋으로 나눈 이유 — 하나로 합치면 어느 쪽을 고쳐야 하는지 알 수 없다.",
              "> 파서 성공률은 *맞는 칸을 집었는가*, 완전 일치율은 *원문이 이미 표준값인가*,",
              "> 정규화 후 일치율은 *규칙까지 태우면 마스터에 넣을 수 있는가* 다.",
              "> **규칙공백 칸이 다음 작업 목록**이고 나머지는 이미 끝난 일이다."]
    L += [""]

    if sc.unscorable_fields:
        L += ["> 채점 제외 — 킷 이름이 스키마에 없어 대조 불가: "
              + ", ".join(f"`{x}`" for x in sc.unscorable_fields), ""]

    gaps = sc.gaps()
    if gaps:
        # 이 절이 다음 작업 목록이다. 규칙이 이미 메우는 칸은 여기 안 나온다.
        L += ["## 🔴 규칙공백 — 다음에 손볼 칸", "",
              "④ Normalize 를 태워도 정답이 안 되는 칸이다. 표기 사전에 없거나 "
              "파서가 다른 값을 집었거나 정답지가 틀렸다.", "",
              "| 문서 | 필드 | 정답 | 파서 원문 | 정규화 후 |", "|---|---|---|---|---|"]
        for c in sorted(gaps, key=lambda x: (x.field_key, x.doc_id)):
            L.append(f"| {c.doc_id} | {c.field_name} | {c.truth} | "
                     f"{c.got if c.got is not None else '—'} | "
                     f"{c.normalized or '—'} |")
        L += [""]

    L += ["## 문서별", "", "| 문서 | 파일 | 상태 |", "|---|---|---|"]
    for d, f, st in sc.docs:
        L.append(f"| {d} | {f} | {st} |")

    L += ["", "## 칸별", "", "| 문서 | 필드 | 정답 | 파서 출력 | 판정 |", "|---|---|---|---|---|"]
    for c in sorted(sc.cells, key=lambda x: (x.doc_id, x.field_key)):
        mark = " ⚠" if c.uncertain else ""
        # 규칙 축이 있으면 판정 뒤에 붙인다 — 표를 하나로 유지한다
        verdict = f"{c.verdict} → {c.rule_state}" if c.rule_state else c.verdict
        L.append(f"| {c.doc_id} | {c.field_name} | {c.truth}{mark} | "
                 f"{c.got if c.got is not None else '—'} | {verdict} |")

    # ── 라벨러별 (2026-08-26) ────────────────────────────────────
    #   골든셋에 AI 초안이 섞여 있다. AI 가 만든 정답으로 AI 를 채점하면 순환이라,
    #   숫자를 하나로 합치면 발표에서 방어할 수 없다. 나눠서 낸다.
    labelers = sc.labelers()
    if len(labelers) > 1:
        L += ["", "## 라벨러별", "",
              "| 라벨러 | 칸 | 정확 | 표기차이 | 정규화대기 | 오답 | 미추출 | 성공률 |",
              "|---|---:|---:|---:|---:|---:|---:|---:|"]
        for who in labelers:
            m = sc.counts(who)
            t = sum(m.values())
            ok = m["정확"] + m["표기차이"] + m["정규화대기"]
            L.append(f"| {who} | {t} | {m['정확']} | {m['표기차이']} | {m['정규화대기']} | "
                     f"{m['오답']} | {m['미추출']} | {ok / t * 100:.0f}% |" if t else
                     f"| {who} | 0 | | | | | | — |")
        human = [w for w in labelers if "AI" not in w]
        if human and len(human) < len(labelers):
            hm = {k: sum(sc.counts(w)[k] for w in human) for k in
                  ("정확", "표기차이", "정규화대기", "오답", "미추출")}
            ht = sum(hm.values())
            hok = hm["정확"] + hm["표기차이"] + hm["정규화대기"]
            L += ["", f"**사람이 검증한 정답만**: {ht}칸 · 성공률 "
                      f"**{hok / ht * 100:.0f}%** · 오답 {hm['오답']}건" if ht else ""]
            L += ["", "> AI 초안은 사람 검증 전이다. 그 행을 기준으로 잰 숫자는 "
                      "**AI 가 만든 정답으로 AI 를 채점**하는 부분이 섞이므로 참고로만 본다.",
                  "> 텍스트 파서는 VLM 과 다른 방식이라 완전한 순환은 아니지만, "
                  "AI 가 놓친 값은 정답지에서도 N/A 로 남아 미추출이 과소평가될 수 있다."]
    return "\n".join(L) + "\n"


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="골든셋 대비 텍스트 파서 채점")
    ap.add_argument("--kit", required=True)
    ap.add_argument("--root", default="raw_file")
    ap.add_argument("--out", default="runs/parser_score.md")
    a = ap.parse_args(argv)

    sc = score(os.path.expanduser(a.kit), a.root)
    md = render(sc)
    os.makedirs(os.path.dirname(os.path.abspath(a.out)) or ".", exist_ok=True)
    with open(a.out, "w", encoding="utf-8") as f:
        f.write(md)
    n = sc.counts()
    print(f"채점 칸 {sum(n.values())} · " + " · ".join(f"{k} {v}" for k, v in n.items()))
    print(f"  {a.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
