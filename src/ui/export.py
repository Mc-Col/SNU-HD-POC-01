# -*- coding: utf-8 -*-
"""엑셀 export — readme/output_sample.xlsx 의 지면을 그대로 따른다.

Output 시트는 전치 구조다(필드가 열, 항목이 행). 마스터DB 적재 대상이므로
열 순서는 계약이다 — schema/fields.yaml 의 정의 순서 = 샘플의 DB CODE 순서.

감사 추적은 별도 시트에 둔다. Output 시트의 모양을 바꾸지 않기 위해서다.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src import schema
from src.contracts import FieldState
from src.ui.source import UiDoc
from src.ui.theme import LABEL

HEAD = Font(bold=True, size=9)
BODY = Font(size=9)
MONO = Font(name="Consolas", size=9)
BAND = PatternFill("solid", fgColor="EAF0F4")
WRAP = Alignment(wrap_text=True, vertical="top")

ROWS = ["DB CODE", "대분류", "분류", "DESCRIPTION", "EXAMPLE",
        "도출 값", "필수여부", "유사 표현"]


def build(d: UiDoc) -> bytes:
    wb = Workbook()
    _output_sheet(wb.active, d)
    _audit_sheet(wb.create_sheet("검증기록"), d)
    _meta_sheet(wb.create_sheet("실행정보"), d)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _output_sheet(ws, d: UiDoc) -> None:
    ws.title = "Output"
    by_key = {r.field_key: r for r in d.records}

    for i, name in enumerate(ROWS, start=2):        # 1행은 샘플과 같이 비워둔다
        c = ws.cell(i, 1, name)
        c.font = HEAD
        c.fill = BAND
    ws.column_dimensions["A"].width = 13

    for j, f in enumerate(schema.all_fields(), start=2):
        ws.column_dimensions[get_column_letter(j)].width = 20
        rec = by_key.get(f.key)
        vals = [
            f.db_code, f.group, f.name, f.desc, f.example,
            (rec.final_value if rec else None),      # 도출 값
            "O" if f.required else None,
            ", ".join(f.aliases) or None,
        ]
        for i, v in enumerate(vals, start=2):
            c = ws.cell(i, j, v)
            c.font = MONO if i == 7 else BODY
            c.alignment = WRAP
    ws.freeze_panes = "B2"


def _audit_sheet(ws, d: UiDoc) -> None:
    cols = ["필드키", "항목", "필수", "안전등급", "상태", "확신도", "임계",
            "원문 라벨", "원문 값", "표준값", "사람 조치", "사람 입력", "최종값",
            "비고", "변환 이력", "근거 위치", "bbox", "재시도"]
    for j, name in enumerate(cols, start=1):
        c = ws.cell(1, j, name)
        c.font = HEAD
        c.fill = BAND
        wide = name in ("비고", "변환 이력")
        ws.column_dimensions[get_column_letter(j)].width = 34 if wide else 15

    for i, r in enumerate(d.records, start=2):
        row = [
            r.field_key, r.field_name, "O" if r.required else "", r.safety,
            LABEL.get(r.state, r.state.value),
            round(r.confidence, 3), r.threshold,
            r.raw_label, r.raw_value, r.value,
            r.human_action or "", r.human_value, r.final_value,
            r.note, " → ".join(r.transform_trace), r.source_locator,
            ",".join(f"{v:g}" for v in r.bbox) if r.bbox else "",
            r.retry_count,
        ]
        for j, v in enumerate(row, start=1):
            c = ws.cell(i, j, v)
            c.font = BODY
            c.alignment = WRAP
    ws.freeze_panes = "C2"


def _meta_sheet(ws, d: UiDoc) -> None:
    res = d.result
    counts = res.counts()
    items: list[tuple[str, object]] = [
        ("문서 ID", res.doc_id),
        ("원본 파일", d.display_name),
        ("데이터 원천", d.origin),
        ("파서 경로", res.records[0].parser.value if res.records else ""),
        ("경로 판정 근거", d.route_reason),
        ("문서 분류", res.triage.document_class.value),
        ("필드 수", len(res.records)),
        ("정상추출", counts.get(FieldState.AUTO.value, 0)),
        ("확인필요", counts.get(FieldState.REVIEW.value, 0)),
        ("N/A", counts.get(FieldState.NA.value, 0)),
        ("필수 전부 해소", "예" if res.approvable else "아니오"),
        ("사람이 고친 필드", sum(1 for r in res.records if r.human_value is not None)),
        ("파이프라인 소요(ms)", res.elapsed_ms),
    ]
    items += [(f"설정 해시 · {k}", v) for k, v in schema.config_hashes().items()]

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 62
    for i, (k, v) in enumerate(items, start=1):
        a = ws.cell(i, 1, k)
        a.font = HEAD
        a.fill = BAND
        b = ws.cell(i, 2, v)
        b.font = BODY
        b.alignment = WRAP


def filename(d: UiDoc) -> str:
    stem = d.display_name.rsplit(".", 1)[0]
    return f"{stem}_master.xlsx"
