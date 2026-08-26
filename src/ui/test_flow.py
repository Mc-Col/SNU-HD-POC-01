# -*- coding: utf-8 -*-
"""화면 자기검증 — 파서가 없어도 흐름과 잠금이 맞는지 확인한다.

    python -m pytest src/ui/test_flow.py -q

여기서 검증하는 것은 '화면이 계약을 따르는가' 하나다.
필수 필드가 남아 있으면 승인이 막히는가, 안전·식별 필드가 정상추출이어도
사람 확인을 요구하는가, 필수여부가 fields.yaml 에서 오는가.
"""
from __future__ import annotations

import io
import os

import pytest
from streamlit.testing.v1 import AppTest

from src import schema
from src.contracts import FieldState
from src.hooks import hooks
from src.ui import approve, export, overlay, pages, screens, source

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
APP = os.path.join(ROOT, "app.py")

screens.STEP_DELAY = 0.0          # 더미 지연은 테스트에서 뺀다


def _to_hitl() -> AppTest:
    at = AppTest.from_file(APP, default_timeout=60)
    at.run()
    at.button(key="btn_fixture").click().run()      # 1 → 3 확인화면
    at.button(key="btn_start").click().run()        # 4 추출 → 5 HITL
    assert at.session_state["stage"] == "hitl"
    return at


# ── 픽스처가 스키마를 이긴다는 착각을 막는다 ──────────────────

def test_meta_comes_from_schema_not_fixture():
    d = source.from_fixture()
    for r in d.records:
        f = schema.get(r.field_key)
        assert r.required == f.required
        assert r.safety == f.safety
        assert r.threshold == f.threshold
        assert r.field_name == f.name


def test_na_records_hold_no_value():
    for r in source.from_fixture().records:
        if r.state is FieldState.NA:
            assert r.value in (None, "", "N/A")
            assert r.note.strip(), "N/A 인데 비고가 없다 — 계약 위반"


# ── 잠금 ──────────────────────────────────────────────────────

def test_submit_locked_until_required_resolved():
    at = _to_hitl()
    d = at.session_state["doc"]

    left = {r.field_key for r in d.unresolved_required}
    # 확신도 미달(REVIEW) · 근거 없음(NA) · 안전·식별 필드(AUTO 이지만 확인 필요)
    assert left == {"engineering_tag_no", "valve_body_material",
                    "actuator_fail_action", "positioner_model_no"}, left
    assert at.button(key="btn_approve").disabled is True

    # 안전·식별 필드는 정상추출이어도 확인 버튼이 나온다
    at.button(key="ok_engineering_tag_no").click().run()
    at.button(key="ok_actuator_fail_action").click().run()
    assert at.button(key="btn_approve").disabled is True

    # 확인필요 → 사람이 값을 고친다
    at.text_input(key="in_valve_body_material").input("WCB").run()
    at.button(key="go_valve_body_material").click().run()

    # N/A → 근거 없음을 사람이 확인한다
    at.button(key="go_positioner_model_no").click().run()

    d = at.session_state["doc"]
    assert d.unresolved_required == []
    assert d.result.approvable is True
    assert at.button(key="btn_approve").disabled is False

    at.button(key="btn_approve").click().run()
    assert at.session_state["stage"] == "done"


def test_human_edit_reaches_excel():
    at = _to_hitl()
    at.text_input(key="in_valve_body_material").input("WCB").run()
    at.button(key="go_valve_body_material").click().run()

    d = at.session_state["doc"]
    rec = d.record("valve_body_material")
    assert rec.human_action == "override"
    assert rec.final_value == "WCB"          # 사람이 고쳤으면 그 값이 최종

    import io

    import openpyxl
    ws = openpyxl.load_workbook(io.BytesIO(export.build(d)))["Output"]
    assert ws.max_column == len(schema.all_fields()) + 1     # A열 + 30필드
    col = next(j for j in range(2, ws.max_column + 1)
               if ws.cell(2, j).value == schema.get("valve_body_material").db_code)
    assert ws.cell(7, col).value == "WCB"                    # 7행 = 도출 값


# ── 일괄 패스 ─────────────────────────────────────────────────

def test_bulk_approve_skips_safety_fields():
    at = _to_hitl()
    at.button(key="btn_bulk").click().run()
    d = at.session_state["doc"]

    for r in d.records:
        if r.state is FieldState.AUTO and r.safety == "normal":
            assert r.human_action == "approve"
        elif r.safety != "normal":
            assert r.human_action is None, "안전·식별 필드가 일괄 승인에 휩쓸렸다"
    assert at.button(key="btn_approve").disabled is True


# ── 근거 ──────────────────────────────────────────────────────

def test_bbox_overlay_renders_only_evidenced_fields():
    from src.ui import overlay
    d = source.from_fixture()
    boxes = overlay.boxes_for(d.records, 1, "valve_body_material")

    keyed = {r.field_key for r in d.records if r.bbox}
    assert len(boxes) == len(keyed) > 0
    assert sum(1 for b in boxes if b[6]) == 1          # 선택된 박스는 하나
    png = overlay.render(d.page_path, 1, boxes)
    assert png[:8] == b"\x89PNG\r\n\x1a\n"

    for r in d.records:
        if r.state is FieldState.NA:
            assert r.bbox is None, "근거 없음인데 지면 위치가 있다"


# ── 확정된 결정이 코드가 아니라 정의서에 있는가 (2026-08-23) ───

def test_rated_cv_is_required_and_in_mvp():
    """Rated CV 는 필수 — 없는 데이터시트는 없다는 도메인 판단.

    2026-08-24 정정: C027 은 `RATED CV MAX` 를 MVP 로 두었으나, 실물 라벨링에서
    벤더가 MAX/NORMAL 을 구분해 적지 않는 경우가 많아 `RATED CV` 하나로 병합했다
    (Metso `Rated Cv: 26`, Fisher `Valve Coefficient: 34.1` 모두 단일값).
    """
    f = schema.get("rated_cv")
    assert f.required and f.mvp
    # 사람이 문서에서 읽는 값이다. 계산은 화면의 [Cv 계산] 버튼이 별도로 한다
    assert "rated_cv_max" not in {x.key for x in schema.all_fields()}
    assert "rated_cv_normal" not in {x.key for x in schema.all_fields()}


def test_required_cv_is_the_process_side_value():
    """Required CV 는 공정이 요구하는 Cv. Rated CV 와 다른 필드다."""
    assert schema.get("required_cv").mvp
    assert schema.get("required_cv").key != schema.get("rated_cv").key


def test_actuator_type_is_not_an_extraction_target():
    """구동부는 FAIL POSITION 만. 따라서 자동 교차검증이 없다."""
    assert not schema.get("actuator_type").mvp
    assert schema.domain_rule("actuator_fail_action")["cross_check"] == []


def test_positioner_stays_required_and_is_resolved_by_human():
    """포지셔너 부재는 규칙으로 필수를 풀지 않고 사람이 N/A 확인한다.

    2026-08-24 정정: `positioner_type` 은 삭제되었다. 데이터시트의
    `Positioner Type` 은 `ELEC. PNEUMATIC` 처럼 전자·공압 구분값이지 모델명이
    아니어서 기준정보로 쓸 수 없다(Metso 실물 확인). 모델명만 남긴다.
    """
    assert schema.get("positioner_model_no").required
    assert "positioner_type" not in {x.key for x in schema.all_fields()}

    at = _to_hitl()
    rec = at.session_state["doc"].record("positioner_model_no")
    assert rec.state is FieldState.NA and not rec.resolved

    at.button(key="go_positioner_model_no").click().run()
    rec = at.session_state["doc"].record("positioner_model_no")
    assert rec.human_action == "na_confirm" and rec.resolved


# ── 자연어 지침 ───────────────────────────────────────────────

@pytest.fixture
def guidance_restored():
    """지침 파일을 건드리는 시험은 원래 내용을 돌려놓는다."""
    orig = io.open(schema.GUIDANCE_PATH, encoding="utf-8").read()
    yield
    io.open(schema.GUIDANCE_PATH, "w", encoding="utf-8", newline="\n").write(orig)
    schema.reload()


def test_guidance_comes_from_yaml():
    g = schema.guidance("positioner_model_no")
    assert g and "N/A" in g["text"]
    assert schema.general_guidance()["text"].strip()
    assert "schema/guidance.yaml" in schema.config_hashes()


def test_guidance_can_be_written_from_the_screen(guidance_restored):
    at = _to_hitl()
    key = "valve_body_material"
    text = "바디와 트림이 한 칸에 있으면 앞이 바디다. (시험 문구)"

    at.text_area(key=f"gd_{key}").input(text).run()
    at.button(key=f"gdsave_{key}").click().run()

    schema.reload()
    assert text in schema.guidance(key)["text"]
    assert hooks.counters["on_rule_edit"] >= 1        # Loop C 입력이 남는다


# ── 쪽 고르기 (화면 3) ────────────────────────────────────────

MULTI = os.path.join(ROOT, "fixtures", "ui", "sample_multipage.pdf")


def _views():
    source.ensure_fixture_page(multi=True)
    return pages.observe.__wrapped__(MULTI, 0.0)      # 캐시 우회


def test_handwritten_mark_is_invisible_to_text_tools():
    """손글씨 폐기 표시는 어느 도구도 못 읽는다 — 비교 화면이 존재하는 이유다."""
    v = {x.page: x for x in _views()}
    assert v[4].date_raw and "1986" in v[4].date_raw
    assert not v[4].superseded and not v[4].marker
    assert v[1].date_raw.startswith("2003")


def test_rule_picks_the_latest_of_two_specs():
    views = _views()
    got, why = pages.rule_would_pick(views, [1, 4], "10FV001")
    assert got == 1 and "2003" in why
    # 후보가 없으면 규칙은 아무거나 고르지 않는다
    assert pages.rule_would_pick(views, [], "10FV001")[0] is None


def test_page_picker_lists_every_page_and_records_the_comparison():
    at = AppTest.from_file(APP, default_timeout=90)
    at.run()
    at.button(key="btn_fixture_multi").click().run()
    assert at.session_state["stage"] == "confirm"
    for n in (1, 2, 3, 4):
        at.checkbox(key=f"pg_{n}")                    # 없으면 KeyError

    at.checkbox(key="pg_1").check().run()
    assert at.session_state["page_candidates"] == [1]
    at.button(key="btn_start").click().run()

    assert at.session_state["stage"] == "hitl"
    assert at.session_state["vlm_page"] == 1
    assert hooks.counters["on_human_action"] >= 1     # 자동 대조가 기록된다


def test_two_candidates_open_the_compare_view():
    at = AppTest.from_file(APP, default_timeout=90)
    at.run()
    at.button(key="btn_fixture_multi").click().run()
    at.checkbox(key="pg_1").check().run()
    at.checkbox(key="pg_4").check().run()

    assert at.session_state["page_candidates"] == [1, 4]
    at.button(key="top_pick_1")   # 실행 버튼은 격자 위에 (스크롤 없이)
    at.button(key="top_pick_4")
    at.button(key="pick_1")       # 비교 화면에도 쪽마다 버튼이 뜬다
    at.button(key="pick_4")
    with pytest.raises(KeyError):
        at.button(key="btn_start")    # 후보가 둘이면 단일 시작 버튼은 없다


# ── 표시원 (화면 5) ───────────────────────────────────────────

def test_flags_are_computed_from_the_shared_assembler():
    """표시는 저장되지 않는다 — 규칙이 바뀌면 다음 렌더에서 바로 달라진다."""
    d = source.from_fixture()
    rec = d.record("valve_body_material")
    got = {f.source for f in d.flags(rec)}
    assert {"확신도", "어휘"} <= got

    # 어휘를 통과하는 값은 표시가 붙지 않는다
    assert not [f for f in d.flags(d.record("actuator_fail_action"))
                if f.source == "어휘"]


def test_bbox_is_normalized_per_contract():
    d = source.from_fixture()
    boxes = overlay.boxes_for(d.records, None, None)
    assert boxes and all(0.0 <= v <= 1.0 for b in boxes for v in b[:4])
    assert overlay.render(d.page_path, 1, boxes)[1:4] == b"PNG"


# ── 사전 승인 (화면 6) ────────────────────────────────────────

def _cand_rows():
    _s, rows = approve.load(approve.FIXTURE)
    return rows


def test_candidates_are_frequency_sorted():
    counts = [r["count"] for r in _cand_rows()]
    assert counts == sorted(counts, reverse=True)


def test_material_fields_offer_no_alias_option():
    """CS → C5 는 탄소강을 크롬몰리강으로 바꿔치기하는 일이다. 선택지에 없어야 한다."""
    at = AppTest.from_file(APP, default_timeout=90)
    at.run()
    at.button(key="btn_approve_open").click().run()
    at.selectbox[0].set_value("합성 픽스처 (시험용)").run()

    mat = at.radio(key="ap_valve_body_material::CS")
    assert approve.ALIAS not in mat.options and mat.value == approve.IGNORE
    maker = at.radio(key="ap_manufacturer::FISHER CONTROLS")
    assert approve.ALIAS in maker.options
    assert maker.value == approve.IGNORE              # 기본값은 무시


def test_approved_yaml_carries_the_seen_line():
    rows = {r["field_key"]: r for r in _cand_rows()}
    picked = [
        {**rows["manufacturer"], "choice": approve.ALIAS, "target": "FISHER"},
        {**rows["valve_seat_material"], "choice": approve.KEEP,
         "target": "SCS14A + STELL."},
    ]
    text = approve.to_yaml(picked, "2026-08-25")

    assert "value_aliases:" in text and "enum_allowed_values:" in text
    assert "to: FISHER" in text or 'to: "FISHER"' in text
    assert "2026-08-25 승인" in text
    assert "원문라벨" in text
    # 재질은 어휘에만 들어가고 값을 바꾸는 규칙이 되지 않는다
    seat = text[text.index("enum_allowed_values:"):]
    assert "valve_seat_material" in seat and "from:" not in seat


def test_screen_never_writes_the_rules_file():
    before = io.open(os.path.join(ROOT, "schema", "rules.yaml"),
                     encoding="utf-8").read()
    rows = {r["field_key"]: r for r in _cand_rows()}
    approve.to_yaml([{**rows["manufacturer"], "choice": approve.KEEP,
                      "target": "FISHER CONTROLS"}], "2026-08-25")
    after = io.open(os.path.join(ROOT, "schema", "rules.yaml"),
                    encoding="utf-8").read()
    assert before == after


# ── 스캔 tif (대상의 71.9%) ───────────────────────────────────

SCAN = os.path.join(ROOT, "fixtures", "ui", "sample_scan.tif")


def test_scan_has_no_badges_and_the_rule_cannot_decide():
    """스캔은 텍스트가 없으니 규칙이 최신성을 못 가린다 — 그래서 사람이 고른다."""
    views = pages.observe.__wrapped__(SCAN, 0.0)
    assert len(views) == 4 and not any(v.has_text for v in views)

    got, why = pages.rule_would_pick(views, [1, 4], None)
    assert got is None and "사람" in why          # 판정불가는 '틀림' 이 아니다


def test_scan_is_rendered_and_boxes_land_on_the_page_we_showed():
    d = source.from_fixture(page_path=SCAN, page=1)
    assert d.page_path.lower().endswith(".png")   # tif 는 PDF 뷰어로 못 띄운다
    assert d.page_no == 1
    assert overlay.boxes_for(d.records, d.page_no, None)

    # 다른 쪽을 떠 왔으면 그 쪽 박스만 그린다 — 1쪽 레코드가 4쪽에 찍히면 거짓이다
    d4 = source.from_fixture(page_path=SCAN, page=4)
    assert d4.page_no == 4
    assert overlay.boxes_for(d4.records, d4.page_no, None) == ()


def test_grid_images_are_shrunk_and_image_formats_share_one_render():
    from PIL import Image
    small = pages.thumbs.__wrapped__(SCAN, 0.0)
    assert small and max(Image.open(small[0]).size) <= pages.GRID_PX * 2
    # 이미지는 dpi 요청이 무시되므로 72·200 이 같은 폴더를 쓴다 (두 번 뜨지 않는다)
    assert pages._bucket(SCAN, 72) == pages._bucket(SCAN, 200) == "orig"
    assert pages._bucket("x.pdf", 72) != pages._bucket("x.pdf", 200)


def test_scan_flows_through_the_page_picker():
    at = AppTest.from_file(APP, default_timeout=120)
    at.run()
    at.session_state["pending_file"] = SCAN
    at.session_state["origin"] = "fixture"
    at.session_state["stage"] = "confirm"
    at.run()

    for n in (1, 2, 3, 4):
        at.checkbox(key=f"pg_{n}")
    at.checkbox(key="pg_2").check().run()          # 도면을 골라도 화면은 막지 않는다
    at.button(key="btn_start").click().run()

    assert at.session_state["stage"] == "hitl"
    assert at.session_state["vlm_page"] == 2
    assert at.session_state["doc"].page_no == 2


# ── 사람이 최종 확정한다 (정책, 2026-08-25) ───────────────────

def test_every_row_offers_a_way_to_change_the_value():
    """AI 판정에 따라 사람의 수정 권한을 제한하지 않는다."""
    at = _to_hitl()
    for rec in at.session_state["doc"].records:
        if rec.state is FieldState.AUTO:
            at.text_input(key=f"ed_{rec.field_key}")    # 없으면 KeyError
        else:
            at.text_input(key=f"in_{rec.field_key}")


def test_confident_value_can_be_corrected_and_stays_countable():
    """자동확정을 사람이 고친 건수가 곧 오적재 관측치다 — state 를 지우면 셀 수 없다."""
    at = _to_hitl()
    key = "manufacturer"                                # 정상추출 · 일반 필드
    rec = at.session_state["doc"].record(key)
    assert rec.state is FieldState.AUTO and rec.human_action is None

    at.text_input(key=f"ed_{key}").input("FISHER CONTROLS").run()
    at.button(key=f"edok_{key}").click().run()

    rec = at.session_state["doc"].record(key)
    assert rec.human_action == "override"
    assert rec.final_value == "FISHER CONTROLS"
    assert rec.state is FieldState.AUTO             # 판정은 지우지 않는다
    assert "수정" in rec.note


def test_safety_field_keeps_both_confirm_and_edit():
    at = _to_hitl()
    at.button(key="ok_actuator_fail_action")         # 확인
    at.text_input(key="ed_actuator_fail_action")     # 수정
    at.text_input(key="ed_engineering_tag_no")


def test_cv_panel_is_reachable_for_the_required_cv_field():
    """필드 이름이 바뀌면(rated_cv_normal → required_cv) 버튼이 조용히 죽는다."""
    at = _to_hitl()
    rec = at.session_state["doc"].record("required_cv")
    assert rec.state is FieldState.NA
    at.number_input(key="cv_q_required_cv")          # 없으면 KeyError


# ── 배선 · 쪽 전달 (서경빈 선임 보고 5-1 · 5-2) ───────────────

class _NoVlm:
    """API 를 부르지 않는 가짜 VLM. 배선만 확인한다."""

    def extract(self, path, triage, fields):
        return []


def test_screen_does_not_assemble_its_own_parser(monkeypatch):
    """배선은 `pipeline.build()` 한 곳에만.

    화면이 손으로 조립하면 DualParser 는 얻어도 **VLM 응답 캐시와 실제
    Normalizer 를 잃는다.** 캐시가 없으면 같은 문서를 두 번 읽을 때 값이
    달라져 철학 6 이 깨진다.
    """
    from types import SimpleNamespace

    from src import pipeline
    from src.parsers.text.adapter import TextParser
    from src.parsers.text.dual import DualParser

    seen = {}

    def spy(**kw):
        seen.update(kw)
        return SimpleNamespace(
            vlm_parser=DualParser(_NoVlm(), TextParser()),
            normalizer=pipeline.DefaultNormalize())

    monkeypatch.setattr(pipeline, "build", spy)
    d = source.from_vlm(os.path.join(ROOT, "fixtures", "ui",
                                     "sample_multipage.pdf"), page=1)

    assert seen, "from_vlm 이 build() 를 부르지 않았다 — 손으로 조립하고 있다"
    assert seen["use_vlm"] is True
    assert "텍스트 대조" in d.route_reason      # 두 경로 대조 결과가 화면에 남는다
    # 텍스트만 나온 후보는 확신도 0 이라 자동확정되지 않는다
    for r in d.records:
        if r.raw_value and r.confidence == 0.0:
            assert r.state is not FieldState.AUTO


def test_pipeline_path_uses_the_page_the_human_picked():
    """쪽 지정 오류는 실측에서 값을 가장 크게 바꾼 결함이다 (d040 46% → 93%)."""
    src_pdf = os.path.join(ROOT, "fixtures", "ui", "sample_multipage.pdf")
    d = source.from_pipeline(src_pdf, use_vlm=False, page=4)

    assert d.page_no == 4
    sel = d.result.triage.selected_page
    assert sel is not None and sel.page == 4
    assert "p4" in d.result.triage.reason

    # 쪽을 주지 않으면 Triage 의 판단을 그대로 쓴다 (덮어쓰지 않는다)
    plain = source.from_pipeline(src_pdf, use_vlm=False)
    assert "사람이 p" not in plain.result.triage.reason


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-q"]))
