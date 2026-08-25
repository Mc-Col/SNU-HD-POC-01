# -*- coding: utf-8 -*-
"""
readme/output_sample.xlsx → schema/fields.yaml 생성

  python tools/gen_schema.py

엑셀이 단일 소스이므로 전사 오류가 없고, 표준 전문가가 엑셀을 고치면 재생성하면 된다.
엑셀에 아직 없는 메타(출처·안전등급·MVP 여부)는 아래 표에서 주입한다.
"""
import os, re, sys, io
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC  = os.path.join(ROOT, "readme", "output_sample.xlsx")
OUT  = os.path.join(ROOT, "schema", "fields.yaml")

# ── 엑셀에 아직 없는 메타 (확정된 결정사항) ──────────────────────────
# 안전등급: safety(안전) / identity(식별) / normal(일반)
SAFETY = {
    "ENGINEERING TAG NO.":   "identity",   # 오적재 시 타 자산 덮어씀
    "ACTUATOR FAIL ACTION":  "safety",     # ATO/ATC 역전, 비상시 거동
}
# MVP 수직 슬라이스 대상 (각각 다른 코드 경로를 대표)
MVP = [
    "ENGINEERING TAG NO.",      # 식별
    "MANUFACTURER",             # 라벨 인접값
    "MODEL NO.",                # 라벨 인접값
    "VALVE BODY SIZE",          # 표기 다양성
    "VALVE BODY RATING",        # 표기 다양성
    "VALVE BODY MATERIAL",      # 자유 텍스트 편차
    "ACTUATOR FAIL ACTION",     # 도메인 규칙 역전
    # 2026-08-24 — RATED CV MAX/NORMAL 두 칼럼이 정격·요구 두 개념으로 분리됨.
    # 위 두 항목의 의도를 새 필드명에 그대로 옮김.
    "RATED CV",                 # 필수 · 밸브 정격 (데이터시트에 항상 있음)
    "REQUIRED CV",              # 선택 · 신규 컬럼 → N/A 경로 + [Cv 계산] 검증
]
# ACTUATOR TYPE 은 27필드에 포함되고 추출도 하지만, MVP 9개에는 넣지 않는다.
# (MVP 는 서로 다른 코드 경로를 대표하는 필드만 고른 것이다.)
# 따라서 FAIL ACTION ↔ ACTUATOR TYPE 교차검증은 MVP 에서 돌릴 수 없고,
# 대신 원문 표기(Fails 유무)와 결과값의 정합성을 검증한다 — src/validate/domain.
# ── 현업 판단 반영 (2026-08-24, 이종수 책임) ──────────────────────
# 라벨링 1건을 실제로 채우면서 드러난 표준-문서 불일치를 정리한 결과.
# 엑셀 원본(output_sample.xlsx)에는 아직 남아 있으므로 여기서 걸러낸다.

DROP = {
    # 문서 라벨 "Positioner Type" 이 표준 "POSITIONER TYPE" 과 글자가 같은데
    # 실제로는 모델번호를 담고 있어 매핑이 반대로 붙는다. 구형 양식은 EP/PP 구분이
    # 없어(당시 전량 PP) Type 칸에 모델명만 적었고, 신형은 표기가 일관되지 않다.
    # → 칼럼을 없애고 MODEL NO. 만 쓴다. 타입은 후속 정비에서 규칙으로 일괄 부여.
    "POSITIONER TYPE",

    # 공정 조건은 모두 Normal 열만 쓴다 (2026-08-24 결정).
    # 이유 둘 — (1) 이미 끝난 엔지니어링 결과를 기록하는 것이지 재검증이 아니다
    #          (2) "항상 Normal 열" 이면 규칙이 하나여서 라벨러·VLM 모두 모호함이 없다.
    #              Min/Nor/Max 중 고르게 하면 손글씨 스캔에서 열 정렬 오독이 생긴다.
    # 최대 압력은 운전 한계 정보로 후속에서 되살릴 가치가 있음 (지금은 손실 없음).
    "MAXIMUM TEMPERATURE",
    "MINIMUM TEMPERATURE",
    "MAXIMUM PRESSURE",
}

# 표준에 없어 새로 추가하는 칼럼 — DB 컬럼 신설이 필요함
ADD = [
    dict(name="NORMAL TEMPERATURE", group="공정 조건", db_code="NEW",
         desc="정상 온도. 공정 조건은 모두 Normal 열을 쓴다",
         required=False, after="NORMAL PRESSURE",
         aliases=["INLET TEMPERATURE, NORMAL", "INLET TEMPERATURE"]),
    # 몸체 형상. 대부분 Globe 지만, 빼면 조회하는 사람이 추측해야 하고
    # Globe(플러그·케이지)와 Ball(볼·시트)은 정비 방법과 예비품이 완전히 다르다.
    # TYPE NAME(용도, derived)과 축이 달라 둘 다 있어야 이름 혼동이 사라진다.
    dict(name="VALVE BODY TYPE", group="밸브 바디", db_code="NEW",
         desc="몸체 형상 (Globe / Angle / Ball 등). 문서에서는 Style 칸의 체크박스",
         required=True, after="VALVE BODY RATING",
         aliases=["STYLE", "VALVE STYLE", "BODY TYPE", "BODY STYLE"]),
    dict(name="REQUIRED CV", group="밸브 성능", db_code="NEW",
         desc="정상 운전 조건에서 요구되는 Cv. 정격 Cv 와 다르며, 공정조건으로 계산 가능",
         required=False, after="RATED CV",
         aliases=["REQ'D FLOW COEFF., CV, NORMAL", "REQ'D FLOW COEFF", "REQUIRED FLOW COEFFICIENT"]),
]

# 정격 Cv 는 밸브 카탈로그 값이라 하나뿐이다(Min/Nor/Max 가 없음).
# 문서의 "Req'd Flow Coeff., Cv" 는 공정 조건에서 계산한 요구 Cv 로 다른 개념이다.
# → RATED CV MAX / NORMAL 두 칼럼을 RATED CV 하나로 합친다.
MERGE = {"RATED CV MAX": "RATED CV"}
MERGE_DROP = {"RATED CV NORMAL"}
MERGE_ALIASES = {
    # ── 구역(section)으로 갈리는 표기 (2026-08-25) ─────────────────────
    # 한 표기를 **여러 필드에 등록**한다. 예전에는 이런 표기를 아예 뺐다 —
    # 한 필드가 이기고 나머지가 굶기 때문이다. 이제 파서가 문서의 구역
    # 이름표를 읽으므로(`src/parsers/text/sections.py`), 밸브 본체 묶음의
    # `Model` 과 포지셔너 묶음의 `Model` 을 갈라 쓸 수 있다.
    #
    # 구역을 못 읽는 문서에서는 후보가 여럿이라 그대로 미매핑이 된다.
    # 즉 되살려도 오답이 늘지 않는다 (`FieldIndex.lookup` 의 마지막 줄).
    "MANUFACTURER": ["Maker"],
    "POSITIONER MANUFACTURER": ["Maker"],
    # `Model No.` 는 MODEL NO. 의 표준명이지만 포지셔너 블록에서도 그 이름을
    # 쓴다 (44LV001 r43 `Model No. | YT-1200`). 표준명 소유 필드가 우선한다는
    # 규칙(`FieldIndex.lookup`)이 있어서 빌려 써도 이름 해석이 깨지지 않는다.
    "POSITIONER MODEL NO.": [
        "Model", "Model #", "Model No.",
        # 구역 접두어가 떨어지면 포지셔너 묶음 한정 표기가 된다.
        # 값이 "모델, 제조사" 두 조각이면 위의 복합 라벨 규칙이 먼저 처리한다.
        "POSITIONER Model No. / Mfr.",
    ],

    # 뺀 것 (2026-08-25 골든셋 대조에서 틀린 값을 만들었다)
    #   "Work Scope 2. 1)" · "Work Scope 2. 3)" : 각주 참조이지 항목명이 아니다
    #                                            ("Note. 3)" 을 뺀 것과 같은 이유)
    #   "Guide Material" → VALVE CAGE MATERIAL : 10FV079 에서 가이드는 케이지와 다른 부품
    #                                            ("SOLID STELLITE" vs 정답 "316 SST")
    #   "End Connection" → VALVE BODY RATING   : 15LV015 에서 연결 형식을 담는다
    #                                            ("RF FLANGED" vs 정답 "600#").
    #                                            "End Connection/Flg. ANSI class" 는 남긴다
    # 킷의 원문라벨에서 수집 (2026-08-25, 골든셋 21건). 사람이 상상해서 채울 수 없는
    # 표기 변종을 실물에서 모은 것이다. 한 표기가 여러 필드에 걸리는 것,
    # 다른 필드의 표준명과 겹치는 것, 복합 라벨 규칙이 이미 처리하는 것은 뺐다.
    "ACTUATOR FAIL ACTION": [
        "Air Fails Valve to", "ACT'N Fail Position", "ACTUATOR Fail Position",
        "ACTUATOR Failure Mode", "Actuator Fail", "Fail",
    ],
    "ACTUATOR TYPE": ["Actuator Style", "Act. Type"],
    "CHARACTERISTIC (TRIM FORM)": [
        "Plug or Cage", "Trim Plug or Cage", "BODY Characteristic", "Flow Char.",
        "Flow Character.", "TRIM Flow Char.", "Valve Body Assembly Characteristic",
    ],
    "ENGINEERING TAG NO.": ["TAG NO", "Tag Number", "Valve Tag #", "Tags"],
    "FLUID NAME": ["GENERAL Fluid", "SERVICE CONDITION Fluid"],
    "FLUID STATE": ["GENERAL Fluid Type"],
    "MODEL NO.": [
        # "Model" · "Model #" 는 포지셔너에도 걸린다 — 구역이 가른다
        "Model", "Model #",
        "Model Number", "VALVE BODY/BONNET Model", "VALVE BODY/BONNET Model No.", "Valve Model",
    ],
    "NORMAL FLOW RATE": [
        "Flow Rate, Give Units", "Liq Flow Rate", "Flow Rate NOR", "Flow Rate Norm.",
        "Volumetric Flow Rate Gas",
    ],
    "NORMAL PRESSURE": [
        "Inlet Pressure", "In.Press. NOR", "Inlet Press", "Inlet Pressure, NOR.", "In.Press.",
        "Inlet Pressure Norm.",
    ],
    "NORMAL TEMPERATURE": [
        "Inlet Temperature, NOR.", "Temp. NOR", "Temperature", "Inlet Temperature Norm.",
        "Temp.",
    ],
    "RATED CV": ["VALVE COEFFICIENT", "Body Rated Cv", "Trim Cv", "Trim Rated Cv", "VALVE TRIM Rated Cv"],
    "REQUIRED CV": [
        "Req'd Flow Coeff., Cv", "Flow Coeff.", "Required Cv NOR.", "Calculated Cv",
        "Calculated Cv NOR", "Calculated Cv Normal", "Required Capacity", "Sizing Coefficient",
    ],
    "VISCOSITY": ["Dynamic Viscosity"],
    "SPECIFIC GRAVITY": ["Density", "SG-MW"],
    "VALVE BODY MATERIAL": [
        # "Material" 단독은 바디 재질로만 둔다. 트림 묶음의 "Material" 은 어느
        # 부품인지 문서가 말해주지 않아 구역으로도 갈리지 않는다
        "Material",
        "Body Material", "Body Matl", "MATERIAL Body", "VALVE BODY / BONNET Material",
        "Body Style Material", "MATERIAL Body/Bonnet",
    ],
    "VALVE BODY RATING": [
        "End Connection/Flg. ANSI class", "Rating", "BODY Rating",
        "VALVE BODY / BONNET Rating", "Ebody Style End Connect", "Pressure Class",
        "VALVE BODY / BONNET Pr. Rating", "Valve Body Assembly Pressure Class",
    ],
    "VALVE BODY SIZE": [
        "Body Size", "VALVE BODY / BONNET Size", "BODY Trim Size", "Size",
        "Valve Body Assembly Size",
    ],
    "VALVE BODY TYPE": ["VALVE BODY / BONNET Type"],
    "VALVE CAGE MATERIAL": [
        "Cage and/or Bushing Material", "Trim Cage and/or Bushing Material", "Cage Material",
        "MATERIAL Cage/Retainer", "MATERIAL Guide Bushing",
        "MATERIAL Retainer or Cage", "Retainer Matl", "Retainer or Cage", "Seat Retainer Mtl",
        "TRIM Retainer / Cage",
    ],
    "VALVE LEAKAGE CLASS": [
        "Shutoff Class", "Leakage Spec.", "Seat Leakage", "BODY Leakage Spec.",
        "VALVE BODY / BONNET Seat Leakage",
    ],
    "VALVE PLUG MATERIAL": [
        "Trim Valve Plug Material", "MATERIAL Plug", "Plug Material", "TRIM Plug Material",
        "MATERIAL Plug/Disc/Ball", "Plug", "Plug Matl", "Plug Mtl",
    ],
    "VALVE SEAT MATERIAL": [
        "Seat Ring Material", "Trim Seat Ring Material", "MATERIAL Seat Ring", "Seat Ring",
        "Seat Ring Mtl", "Soft Seat Matl", "Soft Seat Material", "Trim Disk/Seat Material",
    ],
    "VALVE STEM MATERIAL": [
        "Trim Stem", "MATERIAL Stem", "TRIM Stem Material", "Stem", "Stem Matl", "Stem Mtl",
       
    ],
}

# 문서에 항목 라벨이 없어 다른 값에서 도출해야 하는 필드 (라벨링 3건에서 확인)
# document 로 두면 VLM 이 없는 라벨을 찾다가 N/A 를 낸다.
SOURCE_OVERRIDE = {
    "TYPE NAME":   "derived",   # Tag 접두에서 도출 — FV=Flow, TV=Temp, PV=Press, LV=Level
    "FLUID STATE": "derived",   # FLUID NAME 에서 유추 (유체 사전 필요)
    "MANUFACTURER": "document", # 라벨은 없지만 양식 로고에 있음 → 문서에서 읽는 게 맞다
}

# 필드 설명 보완 — 문서 라벨과 뜻이 어긋나 VLM 이 헷갈리는 것
DESC_OVERRIDE = {
    # 원문 보존 — 벤더마다 조합 규칙이 달라(Fisher 667-EZ, Valstone 1121-880)
    # 분해 사전을 벤더별로 만들면 유지보수 부담이 계속 늘고, 분해하면 되돌릴 수 없다.
    "MODEL NO.": "모델 번호. 문서 표기 그대로 (예: 667-EZ, 1121-880). 구경 표기는 제외",
    "MANUFACTURER": "제조사. 항목 라벨이 아니라 양식의 로고·머리글·꼬리말에 있음",
    "POSITIONER MANUFACTURER": "포지셔너 제조사. 양식 로고·꼬리말에 있을 수 있음",
    "RATED CV": "정격 유량계수 (밸브 카탈로그 값). 요구 Cv 와 다름",
    "TYPE NAME": "밸브 종류. 문서에 항목이 없고 Tag 접두에서 도출 (FV→Flow Control Valve)",
    "FLUID STATE": "유체 상태. 문서에 항목이 없고 FLUID NAME 에서 유추",
}

# 공정 조건은 필수에서 제외 (2026-08-21 결정) — 데이터시트 기재율이 낮음
OPTIONAL_GROUPS = {"공정 조건"}

# 확신도 임계값 기본값 — 개발 30건으로 보정 예정
# 안전·식별 필드는 임계를 1.00 으로 두면 영구 REVIEW 가 되어
# "자동확정 + 사람 확인 표시" 설계와 어긋난다. 임계는 조금 높게 두고,
# 필수 확인은 safety 플래그와 FieldRecord.resolved 가 강제한다.
THRESHOLD = {"safety": 0.95, "identity": 0.95, "normal": 0.90}


def key_of(name: str) -> str:
    k = name.lower()
    k = re.sub(r"\(.*?\)", " ", k)
    k = re.sub(r"[^a-z0-9]+", "_", k).strip("_")
    return k


def q(s):
    """YAML 안전 인용"""
    if s is None:
        return '""'
    s = str(s).strip().replace('"', '\\"')
    return f'"{s}"'


def assert_no_duplicate_alias_keys() -> None:
    """MERGE_ALIASES 에 같은 필드명이 두 번 나오면 멈춘다.

    파이썬 딕셔너리 리터럴은 같은 키가 두 번 나오면 **뒤엣것이 앞엣것을 조용히
    덮는다.** 2026-08-25 에 구역용 표기를 넣었다가 이걸로 잃었다 —
    `"MODEL NO.": ["Model", "Model #"]` 를 위에 적었는데 아래에 이미 같은 키가
    있어서 새로 넣은 표기가 사라졌고, 생성된 fields.yaml 만 봐서는 알 수 없었다.
    """
    import ast

    tree = ast.parse(open(__file__, encoding="utf-8").read())
    for node in ast.walk(tree):
        if not (isinstance(node, ast.Assign)
                and any(getattr(t, "id", "") == "MERGE_ALIASES" for t in node.targets)):
            continue
        names = [k.value for k in node.value.keys if isinstance(k, ast.Constant)]
        dups = {n for n in names if names.count(n) > 1}
        if dups:
            raise SystemExit(f"MERGE_ALIASES 에 중복 키가 있다 (뒤엣것만 살아남는다): {sorted(dups)}")


def main():
    assert_no_duplicate_alias_keys()
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Output"]

    # A열 라벨 → 행 번호
    rows = {}
    for r in range(1, ws.max_row + 1):
        lab = ws.cell(r, 1).value
        if lab:
            rows.setdefault(str(lab).strip(), r)

    R_CODE = rows.get("DB CODE")
    R_GRP  = rows.get("대분류")
    R_NAME = rows.get("분류")
    R_DESC = rows.get("DESCRIPTION")
    R_EX   = rows.get("EXAMPLE")
    R_REQ  = rows.get("필수여부")
    R_ALI  = rows.get("유사 표현")
    if not all([R_CODE, R_GRP, R_NAME, R_DESC, R_REQ]):
        sys.exit(f"[중단] 필수 행을 찾지 못함: {rows}")

    fields, seen = [], set()
    for c in range(2, ws.max_column + 1):
        name = ws.cell(R_NAME, c).value
        if not name:
            continue
        name = str(name).strip()
        if name in DROP or name in MERGE_DROP:
            continue
        name = MERGE.get(name, name)
        grp  = str(ws.cell(R_GRP, c).value or "").strip()
        code = str(ws.cell(R_CODE, c).value or "").strip()

        # 유사 표현: 유사표현 행부터 연속된 행을 모두 수집
        aliases = []
        if R_ALI:
            r = R_ALI
            while r <= ws.max_row:
                lab = ws.cell(r, 1).value
                if r != R_ALI and lab:      # 다음 라벨 행 만나면 종료
                    break
                v = ws.cell(r, c).value
                if v and str(v).strip():
                    aliases.append(str(v).strip())
                r += 1

        req = str(ws.cell(R_REQ, c).value or "").strip().upper() == "O"
        if grp in OPTIONAL_GROUPS:
            req = False                      # 공정 조건 필수 제외 결정 반영

        k = key_of(name)
        if k in seen:
            k = f"{k}_{c}"
        seen.add(k)

        for extra in MERGE_ALIASES.get(name, []):
            if extra not in aliases:
                aliases.append(extra)
        desc = DESC_OVERRIDE.get(name) or str(ws.cell(R_DESC, c).value or "").strip()

        safety = SAFETY.get(name, "normal")
        fields.append(dict(
            key=k, name=name, group=grp, db_code=code,
            desc=desc,
            example=str(ws.cell(R_EX, c).value or "").strip() if R_EX else "",
            required=req, safety=safety, aliases=aliases,
            source=SOURCE_OVERRIDE.get(name, "document"),
            mvp=(name in MVP), threshold=THRESHOLD[safety],
        ))

    # 신규 칼럼 삽입
    for spec in ADD:
        entry = dict(
            key=key_of(spec["name"]), name=spec["name"], group=spec["group"],
            db_code=spec["db_code"], desc=spec["desc"], example="",
            required=spec["required"], safety="normal",
            aliases=list(spec["aliases"]) + [a for a in MERGE_ALIASES.get(spec["name"], [])
                                             if a not in spec["aliases"]],
            mvp=(spec["name"] in MVP), threshold=THRESHOLD["normal"],
        )
        idx = next((i for i, f in enumerate(fields)
                    if f["name"] == spec["after"]), len(fields) - 1)
        fields.insert(idx + 1, entry)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    L = []
    A = L.append
    A("# 컨트롤밸브 기준정보 필드 정의서")
    A("#")
    A("# 이 파일은 readme/output_sample.xlsx 에서 자동 생성됨.")
    A("#   재생성: python tools/gen_schema.py")
    A("# 직접 편집하지 말고 엑셀을 수정한 뒤 재생성할 것.")
    A("#")
    A("# safety : safety(안전) / identity(식별) / normal(일반)")
    A("#          safety·identity 는 확신도와 무관하게 사람 확인이 필요한 필드")
    A("# source : document(문서 추출) / derived(규칙 파생) / system(시스템)")
    A("#          MVP 에서는 전부 document — 스마트 계장 파생 필드는 범위 외")
    A("# mvp    : MVP 수직 슬라이스 대상 여부")
    A("")
    A("meta:")
    A(f"  equipment: {q('Control Valve')}")
    A(f"  field_count: {len(fields)}")
    A(f"  mvp_count: {sum(1 for f in fields if f['mvp'])}")
    A(f"  required_count: {sum(1 for f in fields if f['required'])}")
    A(f"  source_workbook: {q('readme/output_sample.xlsx')}")
    A("")
    A("fields:")
    for f in fields:
        A(f"  - key: {f['key']}")
        A(f"    name: {q(f['name'])}")
        A(f"    group: {q(f['group'])}")
        A(f"    db_code: {q(f['db_code'])}")
        A(f"    desc: {q(f['desc'])}")
        if f["example"]:
            A(f"    example: {q(f['example'])}")
        A(f"    required: {str(f['required']).lower()}")
        A(f"    safety: {f['safety']}")
        A(f"    source: {f.get('source', 'document')}")
        A(f"    mvp: {str(f['mvp']).lower()}")
        A(f"    threshold: {f['threshold']:.2f}")
        if f["aliases"]:
            A("    aliases:")
            for a in f["aliases"]:
                A(f"      - {q(a)}")
        else:
            A("    aliases: []          # 라벨링 중 raw_label 로 수집 예정")
        A("")

    io.open(OUT, "w", encoding="utf-8", newline="\n").write("\n".join(L))

    print(f"생성: {OUT}")
    print(f"  필드 {len(fields)}개 / 필수 {sum(1 for f in fields if f['required'])}개 "
          f"/ MVP {sum(1 for f in fields if f['mvp'])}개")
    print(f"  안전등급: " + ", ".join(
        f"{s}={sum(1 for f in fields if f['safety']==s)}" for s in ("safety","identity","normal")))
    have = sum(1 for f in fields if f["aliases"])
    print(f"  유사표현 보유: {have}/{len(fields)}개 필드  (나머지는 라벨링으로 수집)")
    missing = [f["name"] for f in fields if f["mvp"] and not f["aliases"]]
    if missing:
        print(f"  ※ MVP 필드 중 유사표현 미기재: {len(missing)}개")


if __name__ == "__main__":
    main()
