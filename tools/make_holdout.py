# -*- coding: utf-8 -*-
"""블라인드 홀드아웃 행을 킷에 추가한다.

    python tools/make_holdout.py           뽑을 파일만 보여준다
    python tools/make_holdout.py --write   킷에 빈 행을 추가한다

왜 필요한가
─────────────────────────────────────────────────────────────
지금 골든셋 21건은 **전부 개발 중에 열어 본 문서**다. 규칙이 그 문서들에서
나왔고, 실측된 일반화 격차가 있다(초기 11건 90% ↔ 이후 추가분 78%).
그래서 *"처음 보는 문서에서 몇 %인가"* 를 말할 수 없다.

**소급해서 홀드아웃을 만들 수는 없다.** 새로 라벨해야 한다.

지켜야 할 것 — 이게 전부다
    · 라벨러는 **AI 출력을 보지 않고** 문서만 보고 채운다
    · 개발자(나)는 **그 문서를 열어 보지 않는다.** 렌더도 하지 않는다
    · 개봉은 **1회**다. 결과를 보고 규칙을 고치면 그 순간 홀드아웃이 아니다

왜 고르는 것도 규칙으로 하나
─────────────────────────────────────────────────────────────
사람이 고르면 "읽기 쉬운 것" 이나 "어려운 것" 으로 치우친다. 그래서
**파일명 해시로 뽑는다** — 재현 가능하고 내용과 무관하다. 포맷 비율은
코퍼스 실측(tif 71.9% · pdf 16.6% · xlsx 11.6%)에 맞춘다.
"""
from __future__ import annotations

import argparse
import hashlib
import os
import re
import sys

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

from eval import kit as kitmod                               # noqa: E402
from src import preprocess as pp                             # noqa: E402

# 코퍼스 비율에 맞춘 배분 (합 5)
QUOTA = {"tif": 3, "pdf": 1, "xlsx": 1}
EXCLUDE_CLASS = ("DRAWING", "INSTRUMENT LIST")


def _fmt(name: str) -> str | None:
    e = os.path.splitext(name)[1].lower().lstrip(".")
    if e in ("tif", "tiff"):
        return "tif"
    if e == "pdf":
        return "pdf"
    if e in ("xlsx", "xlsm", "xls"):
        return "xlsx"
    return None


def _rank(name: str) -> str:
    """파일명 해시 — 재현 가능하고 내용과 무관한 순서."""
    return hashlib.sha256(name.encode("utf-8")).hexdigest()


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--kit", default="readme/labeling_kit.xlsx")
    ap.add_argument("--root", default="raw_file")
    ap.add_argument("--write", action="store_true")
    a = ap.parse_args(argv)

    used = {(r.file or "").lower() for r in kitmod.read_kit(a.kit)}
    pool: dict[str, list[str]] = {"tif": [], "pdf": [], "xlsx": []}
    for f in sorted(os.listdir(a.root)):
        if f.lower() in used:
            continue
        k = _fmt(f)
        if not k:
            continue
        # 도면·계기목록은 사양표가 없어 범위 밖이다
        if pp.scope_reason(f):
            continue
        pool[k].append(f)

    picked = []
    for k, n in QUOTA.items():
        cand = sorted(pool[k], key=_rank)
        if len(cand) < n:
            print(f"⚠ {k} 후보 부족: {len(cand)}건", file=sys.stderr)
        picked += cand[:n]

    print(f"제외한 기존 라벨 {len(used)}건 · 후보 "
          f"{ {k: len(v) for k, v in pool.items()} }")
    print("\n뽑은 5건 (파일명 해시 순 — 재현 가능):")
    for f in picked:
        print(f"   {_fmt(f):5s}  {f}")

    if not a.write:
        print("\n`--write` 를 붙이면 킷에 빈 행을 추가한다.")
        print("⚠ 라벨링은 **AI 출력을 보지 않고** 문서만 보고 채울 것.")
        return 0

    import openpyxl
    wb = openpyxl.load_workbook(a.kit)
    ws = wb[kitmod.SHEET]

    # 기존 문서ID 최대값 다음부터
    ids = [str(ws.cell(r, 1).value or "") for r in range(3, ws.max_row + 1)]
    nums = [int(m.group(1)) for i in ids if (m := re.fullmatch(r"d(\d+)", i))]
    nxt = max(nums) + 1 if nums else 1

    # 첫 두 열 위치를 머리글에서 찾는다 (문서ID · 파일명)
    g2 = [str(ws.cell(2, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    col_id = g2.index("문서ID") + 1 if "문서ID" in g2 else 1
    col_file = g2.index("파일명") + 1 if "파일명" in g2 else 2

    row = ws.max_row + 1
    added = []
    for f in picked:
        ws.cell(row, col_id).value = f"d{nxt:03d}"
        ws.cell(row, col_file).value = f
        added.append(f"d{nxt:03d}")
        nxt += 1
        row += 1
    wb.save(a.kit)
    print(f"\n킷에 {len(picked)}행 추가: {', '.join(added)}")
    print("나머지 칸(사양표 페이지 · 28필드)을 문서만 보고 채울 것.")
    print(f"채점할 때는 `--holdout {','.join(added)}` 로 잠가 두고, "
          "**한 번만** 개봉한다.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
