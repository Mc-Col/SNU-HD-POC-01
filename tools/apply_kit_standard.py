# -*- coding: utf-8 -*-
"""라벨링 킷의 정답을 표준값으로 맞춘다.

    python tools/apply_kit_standard.py            차이만 보여준다
    python tools/apply_kit_standard.py --write    실제로 고친다

왜 도구로 하나
─────────────────────────────────────────────────────────────
바꾸는 값은 전부 **규칙 파일이 계산해 주는 것**이다. 손으로 옮기면 25칸에서
오타가 나고, 킷은 계속 자라므로 다음에도 필요하다.

무엇을 바꾸고 무엇을 안 바꾸나
    바꾼다      `value_aliases` 로 표준값이 정해지는 값
    안 바꾼다   허용 어휘 밖의 값 · N/A · 빈칸 · 원문라벨 열
                → 어휘 밖은 사람이 판단할 문제이지 도구가 정할 문제가 아니다

**원문라벨 열은 건드리지 않는다.** 그 열의 존재 이유가 "문서에 뭐라고
적혀 있었나" 이므로 표준화하면 근거가 사라진다.

킷 구조를 아는 곳은 `eval/kit.py` 하나여야 하므로 그쪽 열 매칭을 그대로 쓴다.
`--write` 없이 먼저 돌려 눈으로 확인한다. 되돌리려면 git 이 있다.
엑셀에서 파일을 열어 두면 쓰기가 막힌다 — 닫고 실행할 것.
"""
from __future__ import annotations

import argparse
import os
import sys

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

from eval import compare                                    # noqa: E402
from eval import kit as kitmod                              # noqa: E402
from src import schema                                      # noqa: E402
from src.contracts import ParserType, RawExtraction         # noqa: E402
from src.pipeline import DefaultNormalize                   # noqa: E402

NORM = DefaultNormalize()


def standard(key: str, value: str, label: str = "") -> str | None:
    """규칙 파일이 정하는 표준값. 정해지지 않으면 None.

    **원문라벨을 함께 넘긴다** — 페일액션은 라벨에 따라 결과가 뒤집히므로
    라벨 없이 부르면 그 규칙이 돌지 않는다.
    """
    ex = RawExtraction(field_key=key, raw_value=value, raw_label=label,
                       parser=ParserType.VLM, confidence=1.0)
    out = NORM.run(ex, schema.get(key))[0]
    return None if out is None else str(out)


def _value_columns(ws) -> dict[str, int]:
    """필드 key → 1-기준 열 번호. `eval/kit.py` 와 같은 규칙으로 읽는다."""
    by_name = {schema.norm_label(f.name): f.key for f in schema.all_fields()}
    nc = ws.max_column
    g1 = [ws.cell(1, c).value for c in range(1, nc + 1)]
    g2 = [ws.cell(2, c).value for c in range(1, nc + 1)]
    cur, group = None, []
    for x in g1:
        if x:
            cur = x
        group.append(cur)

    out = {}
    for c in range(nc):
        if str(g2[c] or "").strip() != "정답값":      # 원문라벨은 건드리지 않는다
            continue
        key = by_name.get(schema.norm_label(kitmod._norm_name(group[c])))
        if key:
            out[key] = c + 1
    return out


def _label_columns(ws) -> dict[str, int]:
    """필드 key → 원문라벨 열. 라벨 의존 규칙을 재현하는 데 쓴다."""
    by_name = {schema.norm_label(f.name): f.key for f in schema.all_fields()}
    nc = ws.max_column
    g1 = [ws.cell(1, c).value for c in range(1, nc + 1)]
    g2 = [ws.cell(2, c).value for c in range(1, nc + 1)]
    cur, group = None, []
    for x in g1:
        if x:
            cur = x
        group.append(cur)
    out = {}
    for c in range(nc):
        if str(g2[c] or "").strip() != "원문라벨":
            continue
        key = by_name.get(schema.norm_label(kitmod._norm_name(group[c])))
        if key:
            out[key] = c + 1
    return out


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--kit", default="readme/labeling_kit.xlsx")
    ap.add_argument("--write", action="store_true")
    a = ap.parse_args(argv)

    import openpyxl
    wb = openpyxl.load_workbook(a.kit)
    ws = wb[kitmod.SHEET]
    col = _value_columns(ws)
    print(f"시트 {ws.title!r} · 정답값 열 {len(col)}개")
    if len(col) < 20:
        print("⚠ 열 매칭이 적다 — 킷 구조가 바뀌었을 수 있다.", file=sys.stderr)

    lab = _label_columns(ws)
    changes = []
    for r in range(3, ws.max_row + 1):
        doc = str(ws.cell(r, 1).value or "").strip()
        for key, c in col.items():
            cur = str(ws.cell(r, c).value or "").strip()
            if not cur or compare.is_na(cur):
                continue
            lc = lab.get(key)
            std = standard(key, cur,
                           str(ws.cell(r, lc).value or "").strip() if lc else "")
            if std is not None and std != cur:
                changes.append((r, c, doc, key, cur, std))

    for _r, _c, doc, key, cur, std in changes:
        print(f"   {doc:6s} {key:26s} {cur!r}  →  {std!r}")
    print(f"\n{len(changes)}칸")

    if not a.write:
        print("`--write` 를 붙이면 실제로 고친다. 되돌리려면 git 을 쓴다.")
        return 0
    if not changes:
        return 0

    for r, c, _d, _k, _cur, std in changes:
        ws.cell(r, c).value = std
    wb.save(a.kit)
    print(f"저장: {a.kit}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
