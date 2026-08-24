# -*- coding: utf-8 -*-
"""
라벨링 킷 생성 — schema/fields.yaml → readme/labeling_kit.xlsx

  python tools/gen_labelkit.py             전체 28필드 (MVP 9개는 앞쪽 진한 색)
  python tools/gen_labelkit.py --mvp-only  MVP 8필드만

골든셋 문서의 정답값을 사람이 기입하는 양식이다.
필드 정의(output_sample.xlsx)와는 다른 것 — 이건 "이 문서의 답은 무엇인가"를 적는다.
"""
import os, sys, argparse
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from src import schema

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "readme", "labeling_kit.xlsx")

ROWS = 30                      # 골든셋 기본 30건
NAVY = "1F3064"
HEAD = "E7EDF1"
META = "F2F4F8"
SAFE = "F6EEDF"
FAINT = "F7F8F6"      # 확장 필드 (지금 안 채워도 됨)
FSAFE = "FCF7EE"      # 확장 + 안전
THIN = Side(style="thin", color="C8CEDC")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FORMATS = "xlsx,xlsm,xls,pdf(텍스트),pdf(스캔),tif"
VINTAGE = "현행,레거시"
DOCCLASS = "datasheet,datasheet_embedded,out_of_scope"


# 필드가 이름을 바꾸거나 합쳐졌을 때 기존 입력을 옮기는 표
RENAME = {
    "RATED CV MAX": "RATED CV",         # 정격 Cv 는 하나. Valve Coefficient 값이 여기로
    "REQUIRED CV NORMAL": "REQUIRED CV",  # 킷에서 손으로 붙인 Normal 접미 흡수
    "REQUIRED CV NOR": "REQUIRED CV",
}


def _norm_key(s: str) -> str:
    return " ".join(str(s).upper().split())


def _clean(label) -> str:
    """1행 헤더에서 표준 필드명만 남긴다."""
    s = str(label or "").replace("· ", "").replace("  ※안전", "").strip()
    return RENAME.get(_norm_key(s), s)


def read_existing(path: str) -> dict:
    """기존 킷의 입력을 읽어둔다. 필드가 바뀌어도 이름으로 옮겨 붙일 수 있게.

    반환: {행번호: {"meta": {헤더명: 값}, "fields": {표준필드명: (정답값, 원문라벨)}}}
    """
    if not os.path.exists(path):
        return {}
    try:
        ws = openpyxl.load_workbook(path, data_only=True)["라벨링"]
    except Exception:
        return {}

    # 2행에서 메타 헤더 범위 찾기 — 1행이 "문서 정보" 인 구간
    meta_last = 0
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or "") == "문서 정보":
            meta_last = c
        elif meta_last:
            break

    out = {}
    for r in range(3, ws.max_row + 1):
        meta = {str(ws.cell(2, c).value): ws.cell(r, c).value
                for c in range(1, meta_last + 1)}
        fields = {}
        c = meta_last + 1
        while c < ws.max_column:
            name = _clean(ws.cell(1, c).value)
            if name and name != "비고":
                v, lab = ws.cell(r, c).value, ws.cell(r, c + 1).value
                if v is not None or lab is not None:
                    fields[name] = (v, lab)
            c += 2
        note = ws.cell(r, ws.max_column).value
        if any(v is not None for v in meta.values()) or fields:
            out[r] = {"meta": meta, "fields": fields, "note": note}
    return out


def style_head(c, fill=HEAD, white=False):
    c.font = Font(name="맑은 고딕", size=9, bold=True,
                  color="FFFFFF" if white else "16202B")
    c.fill = PatternFill("solid", fgColor=NAVY if white else fill)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BOX


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--mvp-only", action="store_true", help="MVP 9필드만 (기본은 전체 28필드)")
    ap.add_argument("--rows", type=int, default=ROWS)
    a = ap.parse_args()

    if a.mvp_only:
        fields = list(schema.mvp_fields())
    else:
        # MVP 를 앞에 배치 — 왼쪽부터 채우다 원하는 지점에서 멈출 수 있게
        mvp = list(schema.mvp_fields())
        rest = [f for f in schema.all_fields() if not f.mvp]
        fields = mvp + rest
    n_mvp = sum(1 for f in fields if f.mvp)
    prior = read_existing(OUT)          # 기존 입력 보존
    wb = openpyxl.Workbook()

    # ══════════════ 1) 기입 안내 ══════════════
    ws = wb.active
    ws.title = "기입 안내"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 96

    guide = [
        ("", "골든셋 라벨링 킷", ""),
        ("", "", ""),
        ("", "무엇을 하는 파일인가", "문서 30건을 열어보고 각 필드의 '정답'을 손으로 적는다. "
                                   "AI 정확도를 측정할 유일한 기준이 되므로 AI 출력을 보지 않고 적는다."),
        ("", "왜 AI 출력을 보면 안 되나", "AI 답을 먼저 보면 그것을 무비판 수락하게 되어(앵커링) "
                                        "측정된 정확도가 실제보다 높게 나온다. 정답이 무의미해진다."),
        ("", "", ""),
        ("1", "문서 고르기", "'라벨링' 시트의 파일명 칸에 대상 문서를 적는다. "
                            "포맷·연식이 골고루 섞이도록 아래 배분을 지킨다."),
        ("", "  포맷 배분", "tif(스캔) 15건  ·  pdf 8건  ·  xlsx 5건  ·  xls 1건 + xlsm 1건\n"
                            "※ 실제 대상 909건은 tif 80.7% · pdf 11.9% · xlsx 7.4% 다. 여기서 tif 를 "
                            "절반으로 줄인 것은 의도적이다 — 드문 포맷도 측정되어야 하므로 일부러 "
                            "과대표집한다. 대신 결과를 볼 때 포맷별로 나눠 본다."),
        ("", "  연식 배분", "현행 21건  ·  레거시 9건   ※ 레거시가 전량 30만 태그의 대부분이므로 반드시 포함"),
        ("", "  벤더 배분", "5개사 각 6건 정도. 파일명에 DATA SHEET 가 있는 것에서 고른다"),
        ("", "", ""),
        ("2", "정답값 기입", "문서를 열어 각 필드의 값을 찾아 '정답값' 칸에 적는다."),
        ("", "  ★ 문서에 적힌 대로", "문서에 적힌 값을 그대로 적는다. 문서에 없는 것을 판단해서 "
                                 "만들지 않고, 이상해 보여도 고치지 않는다.\n"
                                 "예) 점도가 0 cP 라고 적혀 있으면 그대로 0 cP 라고 적는다. 물리적으로 "
                                 "불가능해 보여도 그렇다 — 그건 나중에 검증 로직이 화면에 띄워서 "
                                 "엔지니어가 판단할 몫이다. 라벨러가 미리 고치면 그 검증을 측정할 수 없다.\n"
                                 "예) 단위 표기(H · Hr · hr)도 적힌 대로. 평가는 숫자만 대조한다."),
        ("", "  예외 — FAIL ACTION", "이 필드만 표준값으로 적는다: FAIL OPEN 또는 FAIL CLOSE.\n"
                                  "'공기 상실 시 어느 쪽이 되는가' 가 이 필드의 정의이기 때문이다. "
                                  "문서에 적힌 문구는 원문라벨 칸에 남긴다."),
        ("", "  값이 없을 때", "문서에 해당 항목이 아예 없으면  N/A"),
        ("", "  판독이 안 될 때", "항목은 있으나 글씨를 읽을 수 없으면  판독불가"),
        ("", "  확신이 없을 때", "값 앞에 물음표를 붙인다.  예)  ?316SST"),
        ("", "", ""),
        ("2-1", "페이지 기입", "파일이 여러 페이지면 총 몇 장인지, 값을 읽어온 사양표가 몇 번째 페이지인지 "
                             "적는다.  예) 총페이지 8 · 사양표 페이지 4"),
        ("", "  왜 적나", "파이프라인이 '엉뚱한 페이지를 골랐는지' 와 '맞는 페이지에서 잘못 읽었는지' 는 "
                         "완전히 다른 실패다. 이 칸이 없으면 둘을 구분할 수 없다."),
        ("", "  사양표가 여러 장이면", "★ 가장 최신인 페이지를 고른다. 태그가 단독으로 나오는 페이지가 "
                                  "아니다 — 그쪽이 옛 원본일 때가 있다.\n"
                                  "예) 10FV011 은 사양표가 2장인데, p1 은 태그 4개(10-FV-011/012/013/014)를 "
                                  "함께 담은 2003년 Retrofit 사양서이고 p4 는 태그 단독의 1986년 원본이다"
                                  "(수기로 OLD 라고 적혀 있다). 정답은 p1 이다. p4 를 고르면 MODEL NO."
                                  "(657-ED vs 667-ED)와 RATED CV(70.7 vs 95)가 틀린다."),
        ("", "  최신을 어떻게 아나", "① 수기 OLD·SUPERSEDED·폐기 표기가 있으면 그 페이지는 버린다\n"
                                 "② 헤더의 날짜를 본다 (견적번호와 헷갈리지 않게 — 4자리 연도만 날짜다)\n"
                                 "③ RETROFIT / AS-BUILT / REVISED 라고 적힌 쪽이 최신이다\n"
                                 "④ 그래도 못 가리면 비고에 '최신 판정 불가' 라고 적고 넘어간다"),
        ("", "  서로 다른 설비 2대면", "사양표 2장의 태그가 겹치지 않으면 같은 설비의 다른 시점이 아니라 "
                                  "서로 다른 설비다. 이때는 파일명 태그와 같은 쪽을 고른다. 파일명에 태그가 "
                                  "없으면 아무 쪽이나 하나 골라 채우고 비고에 '자산 N건 중 임의 선택' 을 적는다.\n"
                                  "예) 070100_REV0.pdf 는 견적서 한 건에 B10-TV-040 과 B10-TV-1016 의 "
                                  "사양표가 각각 들어 있다."),
        ("", "", ""),
        ("3", "원문 라벨 기입", "문서에 적혀 있던 '항목명'을 그대로 옮겨 적는다. "
                               "표준명과 같으면 비워도 된다."),
        ("", "  왜 중요한가", "이 값들이 모여 유사표현 사전이 된다. 28필드의 표기 변종을 "
                             "머릿속으로 상상해서 채우는 것은 불가능하고, 실물에서 나와야 한다."),
        ("", "  예", "표준명 ACTUATOR FAIL ACTION → 문서에는 'FAIL POSITION' 이라고 적혀 있었다면 "
                     "원문라벨 칸에 FAIL POSITION"),
        ("", "  라벨이 없는 값", "제조사처럼 항목 라벨 없이 양식의 로고·머리글에만 있는 값은 "
                             "위치를 괄호로 적는다.  예)  (좌측 상단 로고)  (하단 꼬리말)\n"
                             "괄호로 시작하면 표기 변종이 아니라 위치 설명으로 처리되어 사전에 들어가지 않는다."),
        ("", "", ""),
        ("4", "문서 분류 기입", "datasheet = 정상 데이터시트  /  "
                              "datasheet_embedded = 정비이력·개조보고서 등에 첨부된 경우  /  "
                              "out_of_scope = 대상 아님"),
        ("", "  왜 적나", "Triage(문서 분류) 정확도를 재는 데 그대로 쓰인다. 추가 비용이 거의 없다."),
        ("", "  정비·개조 보고서도 대상", "REPAIR REPORT · RETROFIT REPORT 도 범위 안이다 "
                                    "(2026-08-24 결정). 안에 사양 페이지가 있으면 datasheet_embedded 로 "
                                    "적고 그 페이지에서 값을 읽는다.\n"
                                    "사양 칸이 대부분 비어 있는 보고서도 있다(정엔지니어링 발행분). 그때는 "
                                    "읽히는 것만 적고 나머지는 N/A 로 둔다 — '이 문서로는 채울 수 없다' 가 "
                                    "측정 대상이다. 억지로 채우지 않는다."),
        ("", "  out_of_scope 로 적는 것", "도면(DRAWING) · 계기 목록(INSTRUMENT LIST) · 견적 색인처럼 "
                                     "설비 사양표가 아예 없는 문서"),
        ("", "", ""),
        ("5", "소요 시간 기입", "문서 1건을 라벨링하는 데 걸린 시간을 분 단위로. "
                              "현행 수기 입력 공수의 실측치가 되므로 대충이라도 적는다."),
        ("", "", ""),
        ("", "어디까지 채우나", "진한 색 9개(MVP)만 채우면 정확도 측정이 된다. 왼쪽부터 배치되어 있다.\n"
                              "흐린 색 19개(· 표시)는 확장용 — 여유가 있을 때 오른쪽으로 이어서 채우면 된다.\n"
                              "나중에 채워도 같은 파일에 이어 쓰면 되므로 킷을 다시 만들 필요가 없다."),
        ("", "", ""),
        ("", "안전 필드 주의", "노란 배경 필드(Tag Number, Actuator Fail Action)는 특히 정확히. "
                             "Tag 를 틀리면 다른 자산의 데이터를 덮어쓰고, Fail Action 은 안전 사양이다."),
        ("", "Fail Action 함정", "양식마다 표기 방식이 두 가지인데, 거의 똑같이 생겼으면서 결과가 정반대다.\n"
                               "  [직접]  Air Fails Valve to : Open   →  FAIL OPEN     공기 상실 시 열림\n"
                               "  [직접]  Air Fails Valve to : Close  →  FAIL CLOSE\n"
                               "  [역전]  Air-to-Open (ATO)           →  FAIL CLOSE    공기로 여는 밸브 = 공기 끊기면 닫힘\n"
                               "  [역전]  Air-to-Close (ATC)          →  FAIL OPEN\n"
                               "\n"
                               "'Air to Open' 과 'Air Fails Valve to Open' 은 결과가 반대다. 문장을 끝까지 읽을 것.\n"
                               "Fails / Fail / 상실 이라는 말이 있으면 [직접], 없으면 [역전]."),
        ("", "체크박스 주의", "Fisher 계열 양식은 Lock / Open / Close 중 하나에 체크(☒)가 되어 있다. "
                            "어느 칸에 표시가 있는지 확인할 것. 손글씨 체크도 있다."),
    ]
    for i, (n, k, v) in enumerate(guide, start=1):
        ws.cell(i, 1, n).font = Font(name="맑은 고딕", size=10, bold=True, color=NAVY)
        c = ws.cell(i, 2, k)
        c.font = Font(name="맑은 고딕", size=(15 if i == 1 else 10),
                      bold=(i == 1 or bool(n) or "주의" in k or "함정" in k),
                      color=NAVY if (i == 1 or bool(n)) else "16202B")
        c.alignment = Alignment(vertical="top")
        d = ws.cell(i, 3, v)
        d.font = Font(name="맑은 고딕", size=10)
        d.alignment = Alignment(vertical="top", wrap_text=True)
        if "함정" in k or "주의" in k:
            for col in (2, 3):
                ws.cell(i, col).fill = PatternFill("solid", fgColor=SAFE)
        if "\n" in v:
            ws.row_dimensions[i].height = 48

    # ══════════════ 2) 라벨링 ══════════════
    ws = wb.create_sheet("라벨링")
    meta_cols = [("문서ID", 10), ("파일명", 30), ("포맷", 13), ("연식", 9),
                 ("문서분류", 20), ("총페이지", 10), ("사양표 페이지", 13),
                 ("라벨러", 10), ("소요(분)", 9)]

    # 1행: 그룹 헤더 / 2행: 실제 헤더
    col = 1
    for name, w in meta_cols:
        ws.column_dimensions[get_column_letter(col)].width = w
        style_head(ws.cell(1, col, "문서 정보"), META)
        style_head(ws.cell(2, col, name), META)
        col += 1

    field_col0 = col
    for f in fields:
        if f.mvp:
            fill = SAFE if f.needs_human else HEAD
        else:
            fill = FSAFE if f.needs_human else FAINT
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        lab = f.name + ("  ※안전" if f.needs_human else "")
        if not f.mvp:
            lab = "· " + lab
        style_head(ws.cell(1, col, lab), fill)
        style_head(ws.cell(1, col + 1), fill)
        style_head(ws.cell(2, col, "정답값"), fill)
        style_head(ws.cell(2, col + 1, "원문라벨"), fill)
        ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions[get_column_letter(col + 1)].width = 16
        col += 2

    ws.column_dimensions[get_column_letter(col)].width = 34
    style_head(ws.cell(1, col, "비고"), META)
    style_head(ws.cell(2, col, "비고"), META)
    last_col = col

    # 데이터 행
    for r in range(3, 3 + a.rows):
        ws.cell(r, 1, f"d{r-2:03d}").font = Font(name="맑은 고딕", size=9, color="6B7681")
        for c in range(1, last_col + 1):
            cell = ws.cell(r, c)
            cell.border = BOX
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(vertical="center")
            if c <= len(meta_cols):
                cell.fill = PatternFill("solid", fgColor="FBFCFB")

    # ── 기존 입력 복원 ──────────────────────────────────
    restored = {"meta": 0, "fields": 0}
    if prior:
        head2 = {str(ws.cell(2, c).value): c for c in range(1, len(meta_cols) + 1)}
        fcol = {}
        c = field_col0
        for f in fields:
            fcol[f.name] = c
            c += 2
        for r, rec in prior.items():
            if r > 2 + a.rows:
                continue
            for k, v in rec["meta"].items():
                if v is not None and k in head2 and k != "문서ID":
                    ws.cell(r, head2[k], v); restored["meta"] += 1
            for name, (v, lab) in rec["fields"].items():
                if name in fcol:
                    if v is not None:
                        ws.cell(r, fcol[name], v); restored["fields"] += 1
                    if lab is not None:
                        ws.cell(r, fcol[name] + 1, lab)
            if rec.get("note") is not None:
                ws.cell(r, last_col, rec["note"])

    # 드롭다운
    def dv(formula, col_idx):
        d = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
        ws.add_data_validation(d)
        L = get_column_letter(col_idx)
        d.add(f"{L}3:{L}{2 + a.rows}")

    dv(FORMATS, 3)
    dv(VINTAGE, 4)
    dv(DOCCLASS, 5)

    ws.freeze_panes = ws.cell(3, field_col0)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    # ══════════════ 3) 필드 참고 ══════════════
    ws = wb.create_sheet("필드 참고")
    heads = ["표준 필드명", "설명", "필수", "안전등급", "예시", "알려진 유사표현"]
    widths = [30, 24, 7, 11, 20, 46]
    for i, (h, w) in enumerate(zip(heads, widths), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        style_head(ws.cell(1, i, h), white=True)
    for r, f in enumerate(fields, start=2):
        vals = [f.name, f.desc, "필수" if f.required else "",
                {"safety": "안전", "identity": "식별"}.get(f.safety, ""),
                f.example, ", ".join(f.aliases)]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(r, i, v)
            c.font = Font(name="맑은 고딕", size=10,
                          bold=(i == 1 and f.needs_human))
            c.alignment = Alignment(vertical="center", wrap_text=(i in (2, 6)))
            c.border = BOX
            if f.needs_human:
                c.fill = PatternFill("solid", fgColor=SAFE)
    ws.freeze_panes = "A2"

    wb.save(OUT)
    print(f"생성: {OUT}")
    print(f"  시트 3개 — 기입 안내 / 라벨링 / 필드 참고")
    print(f"  필드 {len(fields)}개 (MVP {n_mvp} 진한색 + 확장 {len(fields)-n_mvp} 흐린색)"
          f" × (정답값 + 원문라벨) = {len(fields)*2}열")
    print(f"  문서 {a.rows}행  ·  총 {len(meta_cols)+len(fields)*2+1}열")
    print(f"  안전·식별 필드 {sum(1 for f in fields if f.needs_human)}개는 노란 배경")
    print(f"  드롭다운: 포맷 · 연식 · 문서분류")
    if prior:
        print(f"  기존 입력 복원: {len(prior)}행 / 값 {restored['fields']}개")
        lost = set()
        for rec in prior.values():
            lost |= {n for n in rec["fields"] if n not in {f.name for f in fields}}
        if lost:
            print(f"  ※ 삭제된 필드의 입력값은 버려짐: {', '.join(sorted(lost))}")


if __name__ == "__main__":
    main()
